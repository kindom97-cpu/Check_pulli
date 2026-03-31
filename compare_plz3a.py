#!/usr/bin/env python3
"""
Confronto AS-IS vs TO-BE per file PLZ* (PLZ3A, PLZHA, ...)
============================================================
Input supportati:
  - Cartelle sul filesystem  : PLZ*<date>  = AS-IS, PLZ* senza date = TO-BE
  - File zip                 : PLZ*<date>.zip = AS-IS, PLZ*.zip senza date = TO-BE
    (i file nel zip possono stare in una sottocartella, vengono appiattiti)

Join riga-per-riga sulla chiave univoca auto-rilevata, poi confronto
colonna per colonna. Per ogni coppia di file produce un Excel con:

  - Sheet "RIEPILOGO"             : overview tutte le coppie di file
  - Sheet "STRUTTURA_<nome>"      : nomi colonne, tipi, presenza AS-IS/TO-BE
  - Sheet "SINTESI_COL_<nome>"    : per ogni colonna quante righe differiscono
  - Sheet "DIFF_<nome>"           : vista wide — chiave + AS-IS/TO-BE affiancati,
                                    celle gialle dove divergono (max MAX_DIFF_ROWS)
  - Sheet "SOLO_ASIS_<nome>"      : righe presenti solo in AS-IS
  - Sheet "SOLO_TOBE_<nome>"      : righe presenti solo in TO-BE

Uso:
  python3 compare_plz3a.py                          # auto-detect nella CWD
  python3 compare_plz3a.py --base-dir /path/to/dir
  python3 compare_plz3a.py --asis /path/asis --tobe /path/tobe
  python3 compare_plz3a.py --prefix PLZHA           # filtra solo una famiglia
  python3 compare_plz3a.py --output risultati.xlsx
"""

import argparse
import os
import re
import sys
import tempfile
import traceback
import warnings
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=UserWarning)

# ─── Configurazione ───────────────────────────────────────────────────────────
MAX_DIFF_ROWS   = 10_000  # cap righe nel foglio DIFF
CHUNK_INFER     = 200     # righe campione per inferenza tipo
LARGE_FILE_ROWS = 50_000  # oltre questa soglia: confronto a lotti invece di join unico
CHUNK_COMPARE   = 20_000  # dimensione lotto per il confronto su file grandi
MAX_ONLY_ROWS   = 5_000   # cap righe nei fogli SOLO_ASIS / SOLO_TOBE

# ─── Palette colori ──────────────────────────────────────────────────────────
COLOR_HEADER  = "1F4E79"
COLOR_SUBHDR  = "2E75B6"
COLOR_KEY_HDR = "4472C4"
COLOR_DIFF    = "FFE699"
COLOR_ONLY_AS = "F4CCCC"
COLOR_ONLY_TO = "D9EAD3"
COLOR_OK      = "E2EFDA"
COLOR_KO      = "FCE4D6"
COLOR_WARN    = "FFF2CC"


# ─── Helpers stile ───────────────────────────────────────────────────────────
def _fill(c):      return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=11): return Font(bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, row, bg, fg="FFFFFF"):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = _fill(bg); cell.font = _font(bold=True, color=fg)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _row(ws, row, bg=None):
    for cell in ws[row]:
        cell.border = _border(); cell.alignment = Alignment(vertical="center")
        if bg: cell.fill = _fill(bg)

def _autofit(ws, mn=8, mx=50):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, mn), mx)

def _title(ws, text, n_cols):
    ws.append([text])
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws["A1"].fill = _fill(COLOR_HEADER); ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


# ─── Log errori ───────────────────────────────────────────────────────────────
def _log_error(log_path: str, label: str, exc: Exception) -> None:
    """Scrive un errore (con traceback completo) nel file di log."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sep = "=" * 70
    msg = (f"\n{sep}\n"
           f"[{ts}] ERRORE su: {label}\n"
           f"Tipo  : {type(exc).__name__}\n"
           f"Motivo: {exc}\n"
           f"--- Traceback ---\n"
           f"{traceback.format_exc()}"
           f"{sep}\n")
    with open(log_path, "a", encoding="utf-8") as fh:
        fh.write(msg)
    print(f"  [ATTENZIONE]  Errore su {label} — dettagli in: {log_path}")


# ─── Rilevamento famiglie PLZ* ────────────────────────────────────────────────
TS_FOLDER = re.compile(r"\.\d{14}\.\d{14}$")
TS_ZIP    = re.compile(r"\.\d{14}\.\d{14}\.zip$", re.IGNORECASE)

def _has_ts_folder(name): return bool(TS_FOLDER.search(name))
def _has_ts_zip(name):    return bool(TS_ZIP.search(name))

def _plz_prefix(name):
    """Estrae il prefisso logico (es. DW.M.PLZ3A.ALL o DW.M.PLZHA.ALL)."""
    # rimuovi timestamp e .zip
    stem = re.sub(r"\.\d{14}.*$", "", name, flags=re.IGNORECASE)
    stem = re.sub(r"\.zip$", "", stem, flags=re.IGNORECASE)
    return stem.rstrip(".")

def find_pairs(base_dir, prefix_filter=None):
    """
    Scansiona base_dir cercando cartelle e zip PLZ*.
    Quando per lo stesso slot (asis/tobe) esistono sia una cartella che uno zip,
    preferisce lo zip (file originale completo).
    Ritorna dict: { prefix → {"asis": Path, "tobe": Path} }
    """
    base = Path(base_dir)
    # groups[pfx][slot] = list of candidates
    groups = {}

    for p in base.iterdir():
        name_up = p.name.upper()
        if "PLZ" not in name_up:
            continue
        if prefix_filter and prefix_filter.upper() not in name_up:
            continue

        pfx = _plz_prefix(p.name)

        if p.is_dir():
            slot = "asis" if _has_ts_folder(p.name) else "tobe"
        elif p.is_file() and p.suffix.lower() == ".zip":
            slot = "asis" if _has_ts_zip(p.name) else "tobe"
        else:
            continue

        groups.setdefault(pfx, {})
        groups[pfx].setdefault(slot, [])
        groups[pfx][slot].append(p)

    # Per ogni slot scegli il candidato migliore: zip > folder
    result = {}
    for pfx, slots in groups.items():
        result[pfx] = {}
        for slot, candidates in slots.items():
            zips    = [c for c in candidates if c.is_file() and c.suffix.lower() == ".zip"]
            folders = [c for c in candidates if c.is_dir()]
            # preferisci zip se presente, altrimenti folder
            if zips:
                result[pfx][slot] = sorted(zips)[-1]   # prendi il più recente se più zip
            elif folders:
                result[pfx][slot] = sorted(folders)[-1]
    return result


# ─── Estrazione zip → dir temporanea ─────────────────────────────────────────
def extract_zip(zip_path: Path) -> Path:
    """Estrae lo zip in una dir temporanea e ritorna il path alla dir con i CSV."""
    tmp = Path(tempfile.mkdtemp(prefix="plz_compare_"))
    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(tmp)

    # Cerca la cartella che contiene i CSV (appiattimento struttura)
    csv_dirs = set()
    for f in tmp.rglob("*.csv"):
        csv_dirs.add(f.parent)

    if not csv_dirs:
        return tmp

    # Se tutti i CSV sono nella stessa dir, usala
    if len(csv_dirs) == 1:
        return csv_dirs.pop()

    # Altrimenti, copia tutti i CSV in una dir piatta
    flat = tmp / "_flat"
    flat.mkdir()
    for f in tmp.rglob("*.csv"):
        dest = flat / f.name
        if not dest.exists():
            dest.write_bytes(f.read_bytes())
    return flat


# ─── Matching file tra le due cartelle ───────────────────────────────────────
def _logical(filename):
    stem = Path(filename).stem
    cleaned = re.sub(r"\.\d{14}.*$", "", stem)
    return cleaned.rstrip(".")

def match_files(asis_src, tobe_src):
    """
    asis_src / tobe_src possono essere Path (dir o zip).
    Ritorna lista di dict con asis_path, tobe_path, logical_name, short_name.
    """
    def _resolve(src):
        if src is None:
            return None
        src = Path(src)
        if src.is_file() and src.suffix.lower() == ".zip":
            return extract_zip(src)
        return src

    asis_dir = _resolve(asis_src)
    tobe_dir = _resolve(tobe_src)

    def _csvs(d):
        if d is None or not d.exists():
            return {}
        return {_logical(f.name): f for f in d.iterdir() if f.suffix.lower() == ".csv"}

    a_map = _csvs(asis_dir)
    t_map = _csvs(tobe_dir)
    keys  = sorted(set(a_map) | set(t_map))
    return [{"logical_name": k,
             "short_name":   k.split(".")[-1] if "." in k else k,
             "asis_path":    a_map.get(k),
             "tobe_path":    t_map.get(k)} for k in keys]


# ─── Lettura CSV ──────────────────────────────────────────────────────────────
def _has_header(filepath):
    with open(filepath, "r", encoding="utf-8", errors="replace") as fh:
        line = fh.readline().rstrip("\r\n")
    fields = [f.strip() for f in line.split(";") if f.strip()]
    if not fields:
        return False
    return sum(1 for f in fields
               if re.match(r"^[A-Za-z_][A-Za-z0-9_\s]*$", f)) >= max(1, len(fields) * 0.5)

def _dedup_columns(cols):
    """Rinomina colonne duplicate aggiungendo un suffisso _1, _2, ... per evitare
    errori 'Cannot index with multidimensional key' durante il confronto."""
    seen = {}
    result = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            result.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            result.append(c)
    return result

def read_csv(filepath):
    hdr = 0 if _has_header(filepath) else None
    df  = pd.read_csv(filepath, sep=";", header=hdr, dtype=str,
                      encoding="utf-8", encoding_errors="replace",
                      skipinitialspace=True, keep_default_na=False)
    # Deduplicazione nomi colonne prima di qualsiasi operazione di indicizzazione
    df.columns = _dedup_columns([str(c).strip() for c in df.columns])
    # Rimuove SOLO le colonne senza nome (generate dal ';' finale nel CSV).
    # NON rimuovere colonne nominate con valori tutti vuoti: potrebbero essere
    # presenti nell'altro file con dati, creando una falsa differenza strutturale.
    df = df[[c for c in df.columns if c != ""]]
    df = df.apply(lambda c: c.str.strip() if c.dtype == object else c)
    return df


# ─── Inferenza tipo ───────────────────────────────────────────────────────────
def _infer_type(series):
    s = series.dropna().head(CHUNK_INFER)
    if s.empty: return "VUOTO"
    num = pd.to_numeric(s, errors="coerce").notna().mean()
    if num > 0.9:
        return "DECIMALE" if s.str.contains(r"\.", na=False).any() else "INTERO"
    try:
        pd.to_datetime(s, dayfirst=True, errors="raise")
        return "DATA"
    except Exception:
        pass
    return "TESTO"


# ─── Rilevamento chiave ───────────────────────────────────────────────────────
def detect_key(df_a, df_b):
    common = [c for c in df_a.columns if c in df_b.columns]
    if not common: return []
    # singola colonna univoca in entrambi
    for c in common:
        if df_a[c].nunique() == len(df_a) and df_b[c].nunique() == len(df_b):
            return [c]
    # coppia di colonne
    for i, c1 in enumerate(common):
        for c2 in common[i+1:]:
            if (df_a[[c1,c2]].drop_duplicates().shape[0] == len(df_a) and
                df_b[[c1,c2]].drop_duplicates().shape[0] == len(df_b)):
                return [c1, c2]
    return []

def _num_diff(a, b):
    try:
        return float(str(a).replace(",",".")) - float(str(b).replace(",","."))
    except (ValueError, TypeError):
        return ""


# ─── Struttura colonne ────────────────────────────────────────────────────────
def col_structure(df_a, df_b):
    ca, cb = list(df_a.columns), list(df_b.columns)
    all_c  = sorted(set(ca)|set(cb), key=lambda x: (ca.index(x) if x in ca else 9999))
    rows = []
    for c in all_c:
        in_a, in_b = c in ca, c in cb
        ta = _infer_type(df_a[c]) if in_a else "—"
        tb = _infer_type(df_b[c]) if in_b else "—"
        stat = "OK" if (in_a and in_b) else ("SOLO AS-IS" if in_a else "SOLO TO-BE")
        rows.append({
            "COLONNA": c, "IN AS-IS": "Sì" if in_a else "No",
            "IN TO-BE": "Sì" if in_b else "No",
            "POS AS-IS": ca.index(c)+1 if in_a else "—",
            "POS TO-BE": cb.index(c)+1 if in_b else "—",
            "TIPO AS-IS": ta, "TIPO TO-BE": tb,
            "TIPO COERENTE": "Sì" if (in_a and in_b and ta==tb) else ("—" if not (in_a and in_b) else "No"),
            "STATUS": stat,
        })
    return pd.DataFrame(rows)


# ─── Confronto righe ──────────────────────────────────────────────────────────
def compare_rows(df_a, df_b, key_cols):
    common = [c for c in df_a.columns if c in df_b.columns]
    if key_cols:
        merged = pd.merge(df_a[common].reset_index(drop=True),
                          df_b[common].reset_index(drop=True),
                          on=key_cols, how="outer",
                          suffixes=("__AS","__TO"), indicator=True)
        only_as = (merged[merged["_merge"]=="left_only"].drop(columns=["_merge"])
                   .rename(columns=lambda c: c.replace("__AS","").replace("__TO","")))
        only_to = (merged[merged["_merge"]=="right_only"].drop(columns=["_merge"])
                   .rename(columns=lambda c: c.replace("__AS","").replace("__TO","")))
        both    = merged[merged["_merge"]=="both"].drop(columns=["_merge"]).reset_index(drop=True)
    else:
        n = min(len(df_a), len(df_b))
        only_as = df_a.iloc[n:].reset_index(drop=True) if len(df_a)>n else pd.DataFrame(columns=df_a.columns)
        only_to = df_b.iloc[n:].reset_index(drop=True) if len(df_b)>n else pd.DataFrame(columns=df_b.columns)
        left  = df_a[common].iloc[:n].reset_index(drop=True).add_suffix("__AS")
        right = df_b[common].iloc[:n].reset_index(drop=True).add_suffix("__TO")
        both  = pd.concat([left, right], axis=1)
    return both, only_as, only_to

def sintesi_colonne(both, key_cols, value_cols):
    rows = []
    for c in value_cols:
        ca, cb = f"{c}__AS", f"{c}__TO"
        if ca not in both.columns or cb not in both.columns: continue
        eq   = (both[ca].fillna("") == both[cb].fillna("")).sum()
        diff = len(both) - eq
        rows.append({
            "COLONNA": c, "CHIAVE": "Sì" if c in key_cols else "No",
            "RIGHE UGUALI": int(eq), "RIGHE DIVERSE": int(diff),
            "TOT RIGHE": len(both),
            "% DIVERSE": f"{diff/len(both)*100:.1f}%" if len(both) else "0%",
            "STATO": "OK" if diff==0 else "DIFFERENZE",
        })
    return pd.DataFrame(rows)


def compare_rows_chunked(df_a, df_b, key_cols, batch_size=None):
    """
    Confronto per chiave su file grandi tramite lottizzazione di df_b (CHUNK_COMPARE righe per lotto).

    Invece di un unico pd.merge() che materializza l'intero join in memoria, esegue inner-join
    per lotto contro df_a, accumulando contatori e righe diverse senza tenere tutto in RAM.

    Se non è disponibile una chiave, ricade su confronto posizionale a lotti (fallback).

    Ritorna: (both, only_as, only_to, sint, n_diff_total)
      - both          : DataFrame wide (max MAX_DIFF_ROWS) delle righe con almeno una differenza
      - only_as       : DataFrame (max MAX_ONLY_ROWS) righe presenti solo in AS-IS
      - only_to       : DataFrame (max MAX_ONLY_ROWS) righe presenti solo in TO-BE
      - sint          : SINTESI_COL con contatori per colonna
      - n_diff_total  : conteggio totale righe con differenze (non cappato)
    """
    common     = [c for c in df_a.columns if c in df_b.columns]
    value_cols = [c for c in common if c not in key_cols]

    df_a_c = df_a[common].fillna("").reset_index(drop=True)
    df_b_c = df_b[common].fillna("").reset_index(drop=True)

    # ── Fallback posizionale (nessuna chiave rilevata) ────────────────────────
    if not key_cols:
        n = min(len(df_a_c), len(df_b_c))
        sint_rows = []
        for c in common:
            eq   = int((df_a_c[c].iloc[:n].values == df_b_c[c].iloc[:n].values).sum())
            diff = n - eq
            sint_rows.append({
                "COLONNA": c, "CHIAVE": "—",
                "RIGHE UGUALI": eq, "RIGHE DIVERSE": diff,
                "TOT RIGHE": n,
                "% DIVERSE": f"{diff/n*100:.1f}%" if n else "0%",
                "STATO": "OK" if diff == 0 else "DIFFERENZE",
            })
        only_as = df_a_c.iloc[n:].head(MAX_ONLY_ROWS).copy() if len(df_a_c) > n else pd.DataFrame(columns=df_a_c.columns)
        only_to = df_b_c.iloc[n:].head(MAX_ONLY_ROWS).copy() if len(df_b_c) > n else pd.DataFrame(columns=df_b_c.columns)
        return pd.DataFrame(), only_as, only_to, pd.DataFrame(sint_rows), 0

    # ── Confronto per chiave a lotti ──────────────────────────────────────────
    _batch = batch_size if batch_size is not None else CHUNK_COMPARE
    kc_single = key_cols[0] if len(key_cols) == 1 else None

    col_eq_cnt   = {c: 0 for c in value_cols}
    col_diff_cnt = {c: 0 for c in value_cols}
    diff_list    = []   # DataFrame per foglio DIFF (capped a MAX_DIFF_ROWS)
    only_to_list = []   # DataFrame per foglio SOLO_TOBE
    matched_keys = set()
    n_diff_total = 0    # conteggio totale righe con diff (non cappato)

    for start in range(0, len(df_b_c), _batch):
        chunk_b = df_b_c.iloc[start:start + _batch].copy()

        merged = pd.merge(
            df_a_c, chunk_b,
            on=key_cols, how="inner", suffixes=("__AS", "__TO")
        )

        if not merged.empty:
            # Registra chiavi abbinate
            if kc_single:
                matched_keys.update(merged[kc_single].tolist())
            else:
                matched_keys.update(
                    map(tuple, merged[key_cols].itertuples(index=False, name=None))
                )

            # Contatori per colonna (vettorizzati)
            for c in value_cols:
                ca, cb = f"{c}__AS", f"{c}__TO"
                if ca in merged.columns and cb in merged.columns:
                    eq = int((merged[ca] == merged[cb]).sum())
                    col_eq_cnt[c]   += eq
                    col_diff_cnt[c] += len(merged) - eq

            # Maschera righe con almeno una differenza
            diff_mask = pd.Series(False, index=merged.index)
            for c in value_cols:
                ca, cb = f"{c}__AS", f"{c}__TO"
                if ca in merged.columns and cb in merged.columns:
                    diff_mask |= (merged[ca] != merged[cb])

            n_diff_total += int(diff_mask.sum())

            # Accumula righe diverse per il foglio DIFF (fino al cap)
            n_accumulated = sum(len(d) for d in diff_list)
            if n_accumulated < MAX_DIFF_ROWS and diff_mask.any():
                remaining = MAX_DIFF_ROWS - n_accumulated
                diff_list.append(merged[diff_mask].head(remaining))

        # Righe del lotto senza corrispondenza in AS-IS → SOLO_TOBE
        n_only_to = sum(len(d) for d in only_to_list)
        if n_only_to < MAX_ONLY_ROWS:
            if kc_single:
                matched_in_chunk = set(merged[kc_single].tolist()) if not merged.empty else set()
                unmatched = chunk_b[~chunk_b[kc_single].isin(matched_in_chunk)]
            else:
                matched_in_chunk = (
                    set(map(tuple, merged[key_cols].itertuples(index=False, name=None)))
                    if not merged.empty else set()
                )
                unmatched = chunk_b[
                    ~chunk_b[key_cols].apply(tuple, axis=1).isin(matched_in_chunk)
                ]
            remaining = MAX_ONLY_ROWS - n_only_to
            if not unmatched.empty:
                only_to_list.append(unmatched.head(remaining))

    # Righe AS-IS non abbinate a nessuna riga TO-BE → SOLO_ASIS
    if kc_single:
        only_as_df = df_a_c[~df_a_c[kc_single].isin(matched_keys)].head(MAX_ONLY_ROWS).copy()
    else:
        only_as_df = df_a_c[
            ~df_a_c[key_cols].apply(tuple, axis=1).isin(matched_keys)
        ].head(MAX_ONLY_ROWS).copy()

    both       = pd.concat(diff_list,    ignore_index=True) if diff_list    else pd.DataFrame()
    only_to_df = pd.concat(only_to_list, ignore_index=True) if only_to_list else pd.DataFrame()

    # SINTESI_COL (stessa struttura di sintesi_colonne)
    sint_rows = []
    for c in value_cols:
        eq   = col_eq_cnt[c]
        diff = col_diff_cnt[c]
        tot  = eq + diff
        sint_rows.append({
            "COLONNA": c, "CHIAVE": "No",
            "RIGHE UGUALI": eq, "RIGHE DIVERSE": diff,
            "TOT RIGHE": tot,
            "% DIVERSE": f"{diff/tot*100:.1f}%" if tot else "0%",
            "STATO": "OK" if diff == 0 else "DIFFERENZE",
        })
    sint = pd.DataFrame(sint_rows)

    return both, only_as_df, only_to_df, sint, n_diff_total


# ─── Scrittura Excel ──────────────────────────────────────────────────────────
def build_excel(all_families: list[dict], output_path: str, log_path: str = None):
    import openpyxl
    if log_path is None:
        log_path = str(Path(output_path).with_suffix("")) + "_errori.log"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── RIEPILOGO GLOBALE ─────────────────────────────────────────────────────
    ws_r = wb.create_sheet("RIEPILOGO")
    _title(ws_r, "RIEPILOGO CONFRONTO AS-IS vs TO-BE", 11)
    ws_r.append(["FAMIGLIA", "FILE LOGICO", "FILE AS-IS", "FILE TO-BE",
                 "RIGHE AS-IS", "RIGHE TO-BE", "DELTA RIGHE",
                 "COL SOLO AS-IS", "COL SOLO TO-BE",
                 "COLONNE CON DIFF", "RIGHE ABBINATE CON DIFF"])
    _hdr(ws_r, 2, COLOR_SUBHDR)

    all_pair_data = []

    for family in all_families:
        fam_name = family["name"]
        for pair in family["pairs"]:
            a_path, t_path = pair["asis_path"], pair["tobe_path"]
            label = f"{fam_name} / {pair['logical_name']}"
            try:
                df_a = read_csv(a_path) if (a_path and a_path.stat().st_size > 0) else pd.DataFrame()
                df_b = read_csv(t_path) if (t_path and t_path.stat().st_size > 0) else pd.DataFrame()
                n_a, n_b = len(df_a), len(df_b)

                strut = col_structure(df_a, df_b) if (not df_a.empty or not df_b.empty) else pd.DataFrame()
                only_a_cols = strut[strut["STATUS"]=="SOLO AS-IS"]["COLONNA"].tolist() if not strut.empty else []
                only_b_cols = strut[strut["STATUS"]=="SOLO TO-BE"]["COLONNA"].tolist() if not strut.empty else []

                is_large = max(n_a, n_b) > LARGE_FILE_ROWS

                if not df_a.empty and not df_b.empty:
                    key_cols   = detect_key(df_a, df_b)
                    common     = [c for c in df_a.columns if c in df_b.columns]
                    value_cols = [c for c in common if c not in key_cols]
                    if is_large:
                        both, only_as_rows, only_to_rows, sint, n_diff_rows = compare_rows_chunked(df_a, df_b, key_cols)
                        cols_w_diff = sint[sint["STATO"]=="DIFFERENZE"]["COLONNA"].tolist() if not sint.empty else []
                    else:
                        both, only_as_rows, only_to_rows = compare_rows(df_a, df_b, key_cols)
                        sint = sintesi_colonne(both, key_cols, value_cols)
                        cols_w_diff = sint[sint["STATO"]=="DIFFERENZE"]["COLONNA"].tolist() if not sint.empty else []
                        if value_cols and not both.empty:
                            mask = pd.Series([False]*len(both))
                            for c in value_cols:
                                if f"{c}__AS" in both.columns and f"{c}__TO" in both.columns:
                                    mask = mask | (both[f"{c}__AS"].fillna("") != both[f"{c}__TO"].fillna(""))
                            n_diff_rows = int(mask.sum())
                        else:
                            n_diff_rows = 0
                else:
                    is_large=False; key_cols=value_cols=[]; both=pd.DataFrame()
                    only_as_rows=df_a.copy(); only_to_rows=df_b.copy()
                    sint=pd.DataFrame(); cols_w_diff=[]; n_diff_rows=0

                all_pair_data.append({**pair,
                    "family": fam_name, "is_large": is_large,
                    "n_a": n_a, "n_b": n_b,
                    "df_a": df_a, "df_b": df_b,
                    "strut": strut, "sint": sint,
                    "key_cols": key_cols, "value_cols": value_cols, "both": both,
                    "only_as_rows": only_as_rows, "only_to_rows": only_to_rows,
                    "only_a_cols": only_a_cols, "only_b_cols": only_b_cols,
                    "cols_w_diff": cols_w_diff, "n_diff_rows": n_diff_rows,
                    "error": None,
                })

                ok = (not only_a_cols and not only_b_cols and n_diff_rows==0 and n_a==n_b)
                ws_r.append([
                    fam_name, pair["logical_name"],
                    a_path.name if a_path else "— MANCANTE —",
                    t_path.name if t_path else "— MANCANTE —",
                    n_a, n_b, n_a-n_b,
                    ", ".join(only_a_cols) or "—",
                    ", ".join(only_b_cols) or "—",
                    ", ".join(cols_w_diff) or "—",
                    n_diff_rows,
                ])
                _row(ws_r, ws_r.max_row, COLOR_OK if ok else COLOR_KO)

            except Exception as exc:
                _log_error(log_path, label, exc)
                all_pair_data.append({**pair,
                    "family": fam_name, "error": str(exc),
                    "n_a": 0, "n_b": 0,
                })
                ws_r.append([
                    fam_name, pair["logical_name"],
                    a_path.name if a_path else "— MANCANTE —",
                    t_path.name if t_path else "— MANCANTE —",
                    "ERRORE", "ERRORE", "—", "—", "—", "—", str(exc),
                ])
                _row(ws_r, ws_r.max_row, COLOR_WARN)

    _autofit(ws_r); ws_r.freeze_panes = "A3"

    # ── Sheet per ogni coppia ─────────────────────────────────────────────────
    used_snames = {}  # per evitare duplicati tra famiglie diverse

    for pd_ in all_pair_data:
        raw_sname = pd_["short_name"][:19]
        # disambigua se stesso short_name in famiglie diverse
        if raw_sname in used_snames:
            used_snames[raw_sname] += 1
            sname = raw_sname[:17] + str(used_snames[raw_sname])
        else:
            used_snames[raw_sname] = 0
            sname = raw_sname

        lname = pd_["logical_name"]
        fam   = pd_["family"]
        label = f"{fam} / {lname}"

        # Se la coppia ha già dato errore in fase 1, scrivi uno sheet di errore e vai avanti
        if pd_.get("error"):
            ws_err = wb.create_sheet(f"ERRORE_{sname}")
            _title(ws_err, f"[{fam}] ERRORE – {lname}", 1)
            ws_err.append([f"Errore durante l'elaborazione: {pd_['error']}"])
            ws_err.append(["Consultare il file di log per il traceback completo."])
            _row(ws_err, 2, COLOR_WARN)
            continue

        try:
            strut=pd_["strut"]; sint=pd_["sint"]; both=pd_["both"]
            only_as=pd_["only_as_rows"]; only_to=pd_["only_to_rows"]
            key_cols=pd_["key_cols"]; value_cols=pd_["value_cols"]
            is_large=pd_["is_large"]
            n_a=pd_["n_a"]; n_b=pd_["n_b"]; n_diff_rows=pd_["n_diff_rows"]

            # ── STRUTTURA ────────────────────────────────────────────────────────
            ws_s = wb.create_sheet(f"STRUTTURA_{sname}")
            _title(ws_s, f"[{fam}] Struttura colonne – {lname}", 9)
            if strut.empty:
                ws_s.append(["(nessun dato)"])
            else:
                ws_s.append(list(strut.columns)); _hdr(ws_s, 2, COLOR_SUBHDR)
                for _, r in strut.iterrows():
                    ws_s.append(list(r)); rn = ws_s.max_row; _row(ws_s, rn)
                    st=str(r.get("STATUS","")); tc=str(r.get("TIPO COERENTE","Sì"))
                    if   st=="SOLO AS-IS": _row(ws_s, rn, COLOR_ONLY_AS)
                    elif st=="SOLO TO-BE": _row(ws_s, rn, COLOR_ONLY_TO)
                    elif tc=="No":         _row(ws_s, rn, COLOR_WARN)
            _autofit(ws_s); ws_s.freeze_panes = "A3"

            # ── SINTESI COLONNE ───────────────────────────────────────────────────
            ws_sc = wb.create_sheet(f"SINTESI_COL_{sname}")
            _title(ws_sc, f"[{fam}] Sintesi differenze per colonna – {lname}", 7)
            if sint.empty:
                ws_sc.append(["(nessun dato comparabile)"])
            else:
                ws_sc.append(list(sint.columns)); _hdr(ws_sc, 2, COLOR_SUBHDR)
                for _, r in sint.iterrows():
                    ws_sc.append(list(r)); rn=ws_sc.max_row
                    _row(ws_sc, rn, COLOR_KO if str(r.get("STATO",""))=="DIFFERENZE" else COLOR_OK)
            _autofit(ws_sc); ws_sc.freeze_panes = "A3"

            # ── DIFF WIDE ─────────────────────────────────────────────────────────
            ws_d = wb.create_sheet(f"DIFF_{sname}")
            if both.empty or not value_cols:
                _title(ws_d, f"[{fam}] Differenze – {lname}", 1)
                ws_d.append(["(nessuna riga comparabile)"])
            else:
                mask = pd.Series([False]*len(both))
                for c in value_cols:
                    ca, cb = f"{c}__AS", f"{c}__TO"
                    if ca in both.columns and cb in both.columns:
                        mask = mask | (both[ca].fillna("") != both[cb].fillna(""))
                diff_df = both[mask].reset_index(drop=True)
                total_diff = n_diff_rows if is_large else int(mask.sum())
                truncated = total_diff > MAX_DIFF_ROWS
                if len(diff_df) > MAX_DIFF_ROWS:
                    diff_df = diff_df.head(MAX_DIFF_ROWS)

                note = (f" — prime {MAX_DIFF_ROWS:,} su {total_diff:,} righe con diff"
                        if truncated else f" — {len(diff_df):,} righe con diff")
                headers = list(key_cols)
                for c in value_cols:
                    headers += [f"{c} [AS-IS]", f"{c} [TO-BE]", f"DIFF {c}"]

                _title(ws_d, f"[{fam}] Differenze rigaxcolonna – {lname}{note}", len(headers))
                ws_d.append(headers); _hdr(ws_d, 2, COLOR_SUBHDR)
                for idx, k in enumerate(key_cols):
                    ws_d.cell(row=2, column=idx+1).fill = _fill(COLOR_KEY_HDR)

                for _, r in diff_df.iterrows():
                    row_vals = [r.get(k,"") for k in key_cols]
                    for c in value_cols:
                        va = r.get(f"{c}__AS",""); vb = r.get(f"{c}__TO","")
                        row_vals += [va, vb, _num_diff(va, vb)]
                    ws_d.append(row_vals); rn = ws_d.max_row; _row(ws_d, rn)
                    col_offset = len(key_cols) + 1
                    for c in value_cols:
                        if str(r.get(f"{c}__AS","")).strip() != str(r.get(f"{c}__TO","")).strip():
                            ws_d.cell(rn, col_offset).fill     = _fill(COLOR_DIFF)
                            ws_d.cell(rn, col_offset+1).fill   = _fill(COLOR_DIFF)
                        col_offset += 3

            _autofit(ws_d)
            if key_cols:
                ws_d.freeze_panes = f"{get_column_letter(len(key_cols)+1)}3"

            # ── SOLO AS-IS ────────────────────────────────────────────────────────
            ws_a = wb.create_sheet(f"SOLO_ASIS_{sname}")
            nc = max(len(only_as.columns),1)
            only_as_show = only_as.head(MAX_ONLY_ROWS)
            trunc_a = len(only_as) > MAX_ONLY_ROWS
            _title(ws_a, f"[{fam}] Solo in AS-IS – {lname}"
                   + (f" (prime {MAX_ONLY_ROWS:,} di {len(only_as):,})" if trunc_a else ""), nc)
            if only_as_show.empty:
                ws_a.append(["(nessuna riga esclusiva)"])
            else:
                ws_a.append(list(only_as_show.columns)); _hdr(ws_a, 2, COLOR_SUBHDR)
                for _, r in only_as_show.iterrows():
                    ws_a.append(list(r)); _row(ws_a, ws_a.max_row, COLOR_ONLY_AS)
            _autofit(ws_a)

            # ── SOLO TO-BE ────────────────────────────────────────────────────────
            ws_t = wb.create_sheet(f"SOLO_TOBE_{sname}")
            nc = max(len(only_to.columns),1)
            only_to_show = only_to.head(MAX_ONLY_ROWS)
            trunc_t = len(only_to) > MAX_ONLY_ROWS
            _title(ws_t, f"[{fam}] Solo in TO-BE – {lname}"
                   + (f" (prime {MAX_ONLY_ROWS:,} di {len(only_to):,})" if trunc_t else ""), nc)
            if only_to_show.empty:
                ws_t.append(["(nessuna riga esclusiva)"])
            else:
                ws_t.append(list(only_to_show.columns)); _hdr(ws_t, 2, COLOR_SUBHDR)
                for _, r in only_to_show.iterrows():
                    ws_t.append(list(r)); _row(ws_t, ws_t.max_row, COLOR_ONLY_TO)
            _autofit(ws_t)

        except Exception as exc:
            _log_error(log_path, f"{label} (scrittura sheet)", exc)
            ws_err = wb.create_sheet(f"ERRORE_{sname}")
            _title(ws_err, f"[{fam}] ERRORE scrittura sheet – {lname}", 1)
            ws_err.append([f"Errore durante la scrittura: {exc}"])
            ws_err.append(["Consultare il file di log per il traceback completo."])
            _row(ws_err, 2, COLOR_WARN)

    wb.save(output_path)
    print(f"\n[OK] Excel salvato in: {output_path}")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Confronto AS-IS vs TO-BE – PLZ*")
    parser.add_argument("--base-dir", default=".")
    parser.add_argument("--asis",     default=None, help="Path esplicito AS-IS (folder o zip)")
    parser.add_argument("--tobe",     default=None, help="Path esplicito TO-BE (folder o zip)")
    parser.add_argument("--prefix",   default=None, help="Filtra per prefisso (es. PLZHA o PLZ3A)")
    parser.add_argument("--output",   default=None)
    args = parser.parse_args()

    print("=" * 60)
    print("  CONFRONTO AS-IS vs TO-BE – PLZ*")
    print("=" * 60)

    if args.asis or args.tobe:
        # modalità esplicita: una sola famiglia
        families = [{"name": "EXPLICIT",
                     "pairs": match_files(args.asis, args.tobe)}]
    else:
        # auto-detect
        groups = find_pairs(args.base_dir, args.prefix)
        if not groups:
            print("\n[ERRORE] Nessuna cartella/zip PLZ* trovata.")
            sys.exit(1)
        families = []
        for pfx, slots in sorted(groups.items()):
            asis_src = slots.get("asis")
            tobe_src = slots.get("tobe")
            fam_label = pfx.split(".")[-1] if "." in pfx else pfx
            print(f"\n  [{fam_label}]")
            print(f"    AS-IS : {asis_src or '[ATTENZIONE]  NON TROVATA'}")
            print(f"    TO-BE : {tobe_src or '[ATTENZIONE]  NON TROVATA'}")
            pairs = match_files(asis_src, tobe_src)
            if pairs:
                print(f"    File CSV: {len(pairs)}")
                for p in pairs:
                    a = p["asis_path"].name if p["asis_path"] else "MANCANTE"
                    t = p["tobe_path"].name if p["tobe_path"] else "MANCANTE"
                    print(f"      {p['short_name']:30s}  AS-IS: {a}  |  TO-BE: {t}")
                families.append({"name": fam_label, "pairs": pairs})

    if not families:
        print("\n[ERRORE] Nessun file CSV trovato.")
        sys.exit(1)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(args.base_dir, f"confronto_errori_{ts}.log")
    print(f"\n  Elaborazione in corso (max {MAX_DIFF_ROWS:,} righe per foglio DIFF)...")
    print(f"  Log errori    : {log_path}")

    if args.output:
        # output singolo esplicito → tutte le famiglie in un file
        try:
            build_excel(families, args.output, log_path=log_path)
        except Exception as exc:
            _log_error(log_path, args.output, exc)
            print(f"\n[ERRORE] Errore fatale durante la generazione dell'Excel: {exc}")
            print(f"  Dettagli nel log: {log_path}")
            sys.exit(1)
    else:
        # un Excel per ogni singolo file CSV
        for family in families:
            for pair in family["pairs"]:
                sname = pair["short_name"]
                out   = os.path.join(args.base_dir, f"confronto_{sname}_{ts}.xlsx")
                try:
                    build_excel([{"name": family["name"], "pairs": [pair]}], out, log_path=log_path)
                except Exception as exc:
                    _log_error(log_path, out, exc)
                    print(f"  [ERRORE] Skippato {sname} — vedi log per dettagli.")

    print("  Done.")
    if os.path.exists(log_path):
        print(f"  [ATTENZIONE]  Alcuni file hanno generato errori. Consultare: {log_path}")


if __name__ == "__main__":
    main()
