# flowcheck_engine.py
# Motore generale di confronto CSV / ZIP / Cartella
# Produce un Excel per coppia di file (AS-IS vs TO-BE)

from __future__ import annotations

import io
import os
import re
import traceback
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Separatori candidati (ordine decrescente di specificita')
# ---------------------------------------------------------------------------
SEP_CANDIDATES = [";|", ";£", "£|", "|;", "\t", ";", ",", "|", "£"]


# ---------------------------------------------------------------------------
# Rilevamento encoding
# ---------------------------------------------------------------------------

def _detect_encoding(raw_bytes: bytes) -> str:
    """
    Restituisce 'utf-8' se i byte sono UTF-8 valido, altrimenti 'cp1252'.
    cp1252 (Windows-1252) e' il superset di Latin-1 usato dai file DWH
    italiani che contengono caratteri come £ (0xA3) o € (0x80).
    """
    try:
        raw_bytes.decode("utf-8", errors="strict")
        return "utf-8"
    except UnicodeDecodeError:
        return "cp1252"

# ---------------------------------------------------------------------------
# Stili Excel
# ---------------------------------------------------------------------------
_CLR = dict(
    header_bg="1F3864", header_fg="FFFFFF",
    ok_bg="C6EFCE",     ok_fg="276221",
    diff_bg="FFEB9C",   diff_fg="9C5700",
    only_a_bg="FFD7D7", only_a_fg="9C0006",
    only_b_bg="D9E1F2", only_b_fg="1F3864",
    err_bg="FFB3B3",    err_fg="7B0000",
    neutral_bg="F2F2F2",
)

def _fill(hex_bg): return PatternFill("solid", fgColor=hex_bg)
def _font(hex_fg, bold=False): return Font(color=hex_fg, bold=bold)
_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Rilevamento separatore
# ---------------------------------------------------------------------------

def _read_first_lines(filepath: Path, n: int = 30,
                      zip_entry: str | None = None) -> list[str]:
    """
    Legge le prime n righe da:
    - file CSV normale
    - CSV specifico dentro uno ZIP (zip_entry = nome file interno)
    - primo CSV trovato in uno ZIP (zip_entry=None)
    """
    try:
        if filepath.suffix.lower() == ".zip":
            with zipfile.ZipFile(filepath) as zf:
                target = zip_entry
                if target is None:
                    csv_names = [x for x in zf.namelist()
                                 if x.lower().endswith(".csv") and not x.startswith("__")]
                    target = csv_names[0] if csv_names else None
                if target:
                    with zf.open(target) as fh:
                        raw_bytes = fh.read(16384)
                        enc = _detect_encoding(raw_bytes)
                        raw = raw_bytes.decode(enc, errors="replace")
                        return raw.splitlines()[:n]
        else:
            with open(filepath, "rb") as fh:
                raw_bytes = fh.read(16384)
            enc = _detect_encoding(raw_bytes)
            with open(filepath, encoding=enc, errors="replace") as fh:
                return [fh.readline() for _ in range(n)]
    except Exception:
        pass
    return []


def _build_sep_candidates(lines: list[str]) -> list[str]:
    """
    Costruisce la lista di separatori candidati unendo quelli statici con
    eventuali separatori compositi ;X rilevati dinamicamente nelle righe.
    Questo permette di gestire qualsiasi variante (;#, ;£, ;|, ;@, ...) senza
    dover aggiornare la lista fissa.
    """
    import collections

    cands = list(SEP_CANDIDATES)

    # Cerca il carattere X che segue ';' in modo costante su tutte le righe
    after_semi: collections.Counter = collections.Counter()
    non_empty_lines = [l for l in lines if l.strip()]
    for line in non_empty_lines:
        for i in range(len(line) - 1):
            if line[i] == ";":
                nxt = line[i + 1]
                # ignora spazi, virgolette, newline, alfanumerici, altro ;
                if not nxt.isalnum() and nxt not in (";", " ", "\n", "\r", '"', "'"):
                    after_semi[nxt] += 1

    if after_semi:
        best_char, count = after_semi.most_common(1)[0]
        # accettiamo il candidato se appare in almeno il 50% delle righe non vuote
        threshold = max(1, len(non_empty_lines) * 0.5)
        if count >= threshold:
            # Controllo aggiuntivo: ';X' e' un VERO separatore composto solo se
            # OGNI ';' nel file e' seguito da X.
            # Se alcuni ';' non sono seguiti da X (es. ultimo campo senza prefisso),
            # allora X e' un prefisso-valore, non parte del separatore.
            # In quel caso NON inseriamo il composto e usiamo solo ';'.
            total_semi = sum(line.count(';') for line in non_empty_lines)
            if count >= total_semi:
                compound = f";{best_char}"
                if compound not in cands:
                    cands.insert(0, compound)   # massima priorita'

    return cands


def detect_separator(filepath: str | Path,
                     candidates: list[str] | None = None,
                     zip_entry: str | None = None) -> str:
    """
    Rileva il separatore CSV piu' probabile usando un punteggio
    avg_occorrenze / (1 + varianza) su max 30 righe.

    filepath  : percorso al file CSV o ZIP
    candidates: lista separatori da testare; None = auto (statici + dinamici)
    zip_entry : nome del file CSV dentro lo ZIP (None = primo trovato)
    """
    import statistics

    filepath = Path(filepath)
    lines = _read_first_lines(filepath, n=30, zip_entry=zip_entry)
    if not lines:
        return ";"

    cands = list(candidates) if candidates else _build_sep_candidates(lines)

    best_sep, best_score = ";", -1.0
    for sep in cands:
        counts = [line.count(sep) for line in lines if line.strip()]
        if not counts or max(counts) == 0:
            continue
        avg = statistics.mean(counts)
        var = statistics.variance(counts) if len(counts) > 1 else 0
        score = avg / (1.0 + var)
        if score > best_score:
            best_score, best_sep = score, sep
    return best_sep


# ---------------------------------------------------------------------------
# Lettura CSV (con auto-detect del separatore)
# ---------------------------------------------------------------------------

def _clean_str_series(s: pd.Series) -> pd.Series:
    """
    Normalizza una Series di stringhe per la lettura e il confronto:
    - strip leading/trailing whitespace ASCII
    - rimuove spazi non-breaking (U+00A0) da bordi
    - collassa spazi interni multipli in uno singolo
    """
    return (
        s.str.strip()
         .str.replace("\xa0", " ", regex=False)   # non-breaking space -> spazio
         .str.replace(r"[ \t]+", " ", regex=True)  # spazi/tab multipli -> singolo
         .str.strip()
    )


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    """Applica _clean_str_series a tutte le colonne stringa del DataFrame.
    Usa is_string_dtype per compatibilità con pandas 2.0+ (StringDtype != object)."""
    return df.apply(
        lambda col: _clean_str_series(col)
        if pd.api.types.is_string_dtype(col) or col.dtype == object
        else col
    )


def _strip_sep_prefixes(df: pd.DataFrame, sep: str) -> pd.DataFrame:
    """
    Rimuove il prefisso del separatore/valore dai nomi colonna e dai valori stringa.

    Contesti gestiti:
    A) sep=';£'  (PLZCA): primo campo mantiene '£' → '£NUM_CONTR' → 'NUM_CONTR'
    B) sep=';'   (ABK001FW): ogni valore ha prefisso '£' → '£NUM_CONTR' → 'NUM_CONTR'
    C) sep=';#'  (TO-BE):   '#MITT' → 'MITT'

    Strategia in due fasi:
    1. Ricava i caratteri speciali del separatore (parte non-alfanumerica).
    2. Se dopo lo strip dal separatore i nomi colonna o i valori hanno ancora
       un prefisso non-alfanumerico noto (£ # ; | , @ ~), lo aggiunge al set
       di caratteri da eliminare (prefisso-valore indipendente dal separatore).
    """
    _KNOWN_VALUE_PREFIXES = frozenset('£#;|,@~')

    # 1. Caratteri speciali del separatore
    strip_set = {c for c in sep if not (c.isalnum() or c == '_')}

    # 2. Rileva prefisso-valore dai nomi colonna
    #    (copre il caso sep=';' dove '£' NON è nel separatore)
    for col in df.columns:
        if col and col[0] in _KNOWN_VALUE_PREFIXES and col[0] not in strip_set:
            candidate = col[0]
            n_with = sum(1 for c in df.columns if c.startswith(candidate))
            # Accettiamo se almeno il 30% delle colonne porta il prefisso
            if n_with >= max(1, len(df.columns) * 0.30):
                strip_set.add(candidate)
            break

    # 3. Rileva prefisso-valore dai dati della prima colonna (fallback)
    if df.columns.size > 0:
        sample = df[df.columns[0]].dropna().head(30)
        for val in sample:
            if isinstance(val, str) and len(val) > 1 and val[0] in _KNOWN_VALUE_PREFIXES:
                candidate = val[0]
                if candidate not in strip_set:
                    n_match = sum(1 for v in sample
                                  if isinstance(v, str) and v.startswith(candidate))
                    if n_match >= len(sample) * 0.50:
                        strip_set.add(candidate)
                break

    if not strip_set:
        return df

    strip_chars = ''.join(sorted(strip_set))

    # Strip dai nomi colonna
    new_cols = [c.lstrip(strip_chars) for c in df.columns]
    new_cols = _dedup_columns(new_cols)

    df = df.copy()
    df.columns = new_cols

    # Strip dai valori stringa (solo lstrip: non tocca caratteri interni).
    # pd.api.types.is_string_dtype cattura sia 'object' che StringDtype (pandas 2.0+).
    for col in df.columns:
        if pd.api.types.is_string_dtype(df[col]):
            df[col] = df[col].str.lstrip(strip_chars)

    return df


def _dedup_columns(cols: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            result.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            result.append(c)
    return result


def _has_header_from_text(text: str, sep: str) -> bool:
    """
    Euristica robusta per rilevare se la prima riga e' un header o gia' un dato.

    Logica:
    - I nomi colonna DB/DWH tipici matchano ^[A-Za-z_][A-Za-z0-9_]*$
      (es. C_FILIALE, N_CONTRATTO, D_DECORRENZA).
    - Le righe dati contengono di solito ID numerici, date (tutte cifre),
      importi, ragioni sociali con spazi — che NON matchano quel pattern.

    Regole:
    1. Se almeno il 50 % dei valori non-vuoti matchano il pattern colonna
       E almeno uno contiene '_' -> header quasi certo.
    2. Se almeno il 75 % matchano (anche senza '_') -> header.
    3. Altrimenti -> dati, nessun header.
    """
    lines = [ln.rstrip("\r\n") for ln in text.splitlines() if ln.strip()]
    if not lines:
        return True

    first_vals = [v.strip().strip('"').strip("'") for v in lines[0].split(sep)]
    non_empty = [v for v in first_vals if v]
    if not non_empty:
        return True

    # Rimuovi prefissi-valore noti (£, #, ;, |, …) prima del test sul pattern.
    # Questo gestisce file DWH dove i nomi colonna hanno prefisso separatore
    # (es. '£NUM_CONTR' → 'NUM_CONTR', '#MITT' → 'MITT').
    _PFXS = '£#;|,@~'
    cleaned = [v.lstrip(_PFXS) for v in non_empty]

    col_pat = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
    col_like = [v for v in cleaned if col_pat.match(v) and not v.isdigit()]
    ratio = len(col_like) / len(non_empty)
    has_underscore = any("_" in v for v in col_like)

    if ratio >= 0.50 and has_underscore:
        return True
    if ratio >= 0.75:
        return True
    return False


def _has_header(filepath: str | Path, sep: str = ";") -> bool:
    """Legge le prime righe del file (con encoding auto-detect) e delega a _has_header_from_text."""
    try:
        with open(filepath, "rb") as fh:
            raw_bytes = fh.read(16384)
        enc = _detect_encoding(raw_bytes)
        text = raw_bytes.decode(enc, errors="replace")
        return _has_header_from_text(text, sep)
    except Exception:
        return True                        # in caso di errore assume header


def read_csv(filepath: str | Path, sep: str | None = None) -> pd.DataFrame:
    """
    Legge un CSV (anche da dentro un ZIP se filepath e' un Path di ZIP).
    sep=None => auto-detect.
    """
    filepath = Path(filepath)
    if sep is None:
        sep = detect_separator(filepath)

    # engine python obbligatorio per separatori multi-carattere
    engine = "python" if len(sep) > 1 else "c"
    sep_param = re.escape(sep) if engine == "python" else sep

    hdr = 0 if _has_header(filepath, sep) else None

    # Rileva encoding prima di passarlo a pandas
    with open(filepath, "rb") as fh:
        enc = _detect_encoding(fh.read(16384))

    df = pd.read_csv(
        filepath,
        sep=sep_param,
        header=hdr,
        dtype=str,
        encoding=enc,
        encoding_errors="replace",
        skipinitialspace=True,
        keep_default_na=False,
        on_bad_lines="skip",
        engine=engine,
        index_col=False,   # evita che pandas usi col 0 come indice quando i dati
                           # hanno un campo più dell'intestazione (trailing sep)
    )
    # Quando non c'e' header pandas assegna interi (0, 1, 2…);
    # li rinominiamo in Col_1, Col_2, … prima di ogni altro processing.
    if hdr is None:
        df.columns = [f"Col_{i + 1}" for i in range(len(df.columns))]
    df.columns = _dedup_columns([str(c).strip() for c in df.columns])
    # Rimuovi prefissi separatore da nomi colonna e valori (fix shift prima colonna)
    df = _strip_sep_prefixes(df, sep)
    # Rimuovi colonne-artefatto: nome vuoto o solo caratteri speciali (es. '#', '£')
    df = df[[c for c in df.columns if not _is_artifact_col(c)]]
    df = _clean_df(df)
    return df


def _is_artifact_col(col_name: str) -> bool:
    """
    True se il nome colonna e' un artefatto del separatore CSV e va scartato:
    - stringa vuota           (es. trailing ';' -> '')
    - solo caratteri speciali (es. '#', '£', '|', ';#', ';£')
    - auto-nome pandas        (es. 'Unnamed: 33') generato da trailing sep
      quando l'header ha un separatore finale ma i dati no
    """
    stripped = col_name.strip()
    if not stripped:
        return True
    # nessun carattere \w (lettera, cifra, _) -> artefatto
    if not re.search(r"\w", stripped):
        return True
    # pandas auto-naming per colonne senza nome: 'Unnamed: N'
    if re.match(r"^Unnamed:\s*\d+$", stripped, re.IGNORECASE):
        return True
    return False


def read_csv_from_zip(zip_path: str | Path, csv_name: str, sep: str | None = None) -> pd.DataFrame:
    """Estrae un CSV da uno ZIP in memoria e lo legge."""
    zip_path = Path(zip_path)
    if sep is None:
        # Rileva il separatore dal CSV specifico, non dal primo dello ZIP
        sep = detect_separator(zip_path, zip_entry=csv_name)
    engine = "python" if len(sep) > 1 else "c"
    sep_param = re.escape(sep) if engine == "python" else sep

    with zipfile.ZipFile(zip_path) as zf:
        with zf.open(csv_name) as fh:
            raw_bytes = fh.read()
    enc = _detect_encoding(raw_bytes)
    raw = raw_bytes.decode(enc, errors="replace")

    hdr_flag = _has_header_from_text(raw, sep)
    hdr = 0 if hdr_flag else None
    df = pd.read_csv(
        io.StringIO(raw),
        sep=sep_param,
        header=hdr,
        dtype=str,
        encoding_errors="replace",
        skipinitialspace=True,
        keep_default_na=False,
        on_bad_lines="skip",
        engine=engine,
        index_col=False,   # evita che pandas usi col 0 come indice quando i dati
                           # hanno un campo più dell'intestazione (trailing sep)
    )
    # Quando non c'e' header pandas assegna interi (0, 1, 2…);
    # li rinominiamo in Col_1, Col_2, … prima di ogni altro processing.
    if hdr is None:
        df.columns = [f"Col_{i + 1}" for i in range(len(df.columns))]
    df.columns = _dedup_columns([str(c).strip() for c in df.columns])
    # Rimuovi prefissi separatore da nomi colonna e valori (fix shift prima colonna)
    df = _strip_sep_prefixes(df, sep)
    # Rimuovi colonne-artefatto: nome vuoto o solo caratteri speciali (es. '#', '£')
    df = df[[c for c in df.columns if not _is_artifact_col(c)]]
    df = _clean_df(df)
    return df




# ---------------------------------------------------------------------------
# Matching file tra AS-IS e TO-BE
# ---------------------------------------------------------------------------

def _stem_key(name: str) -> str:
    """
    Chiave di matching normalizzata: rimuove timestamp, estensione e
    prefissi tecnici noti (DW, D, M, PLZxx), conservando il nome tabella
    anche quando composto da piu' parti (es. ABK001FW_DANNI, ABBINAMENTO_PLZ_DANNI).

    Esempi:
      DW.D.PLZHA.AVTBCODI.20260326000000.000L.20260326030157.csv  => AVTBCODI
      DW.D.PLZ3A.ABBINAMENTO_PLZ_DANNI.20260326000000.000L.csv    => ABBINAMENTO_PLZ_DANNI
      DW.D.PLZAA.ABK001FW_DANNI.csv                               => ABK001FW_DANNI
      DW.D.PLZBA.STORNI_DANNI.csv                                 => STORNI_DANNI
    """
    name = Path(name).stem.upper()
    # Rimuovi blocchi timestamp (es. 20260326000000) e suffissi tipo 000L
    name = re.sub(r"\b\d{14}\b", "", name)
    name = re.sub(r"\b\d{3}L\b", "", name)
    # Normalizza separatori in underscore, pulisci underscore multipli
    name = re.sub(r"[.\-]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    # Rimuovi prefissi tecnici noti (DW_D_PLZxx_, DW_M_PLZxx_, D_PLZxx_, ecc.)
    # PLZxx = PLZ seguito da esattamente 2 caratteri alfanumerici (es. 3A, HA, AA ...)
    name = re.sub(r"^(DW_)?[DM]_PLZ\w{2}_", "", name)
    # Rimuovi eventuali residui DW_D_ / DW_M_ / D_ / M_ rimasti
    name = re.sub(r"^(DW_)?[DM]_", "", name)
    name = name.strip("_")
    return name if name else Path(name).stem.upper()


def match_files(
    asis_sources: list[tuple[str, str]],   # (display_name, path_or_zipentry)
    tobe_sources: list[tuple[str, str]],
) -> list[dict]:
    """
    Data due liste di (label, path), restituisce coppie abbinate
    [{'label': str, 'asis': str, 'tobe': str}, ...].
    """
    tobe_by_key: dict[str, tuple[str, str]] = {}
    for label, path in tobe_sources:
        tobe_by_key[_stem_key(label)] = (label, path)

    pairs = []
    for label, path in asis_sources:
        key = _stem_key(label)
        if key in tobe_by_key:
            tb_label, tb_path = tobe_by_key[key]
            pairs.append({"label": key, "asis_label": label, "tobe_label": tb_label,
                          "asis_path": path, "tobe_path": tb_path})
        else:
            pairs.append({"label": key, "asis_label": label, "tobe_label": None,
                          "asis_path": path, "tobe_path": None})
    return pairs


# ---------------------------------------------------------------------------
# Detect join key
# ---------------------------------------------------------------------------

def detect_join_key(df_a: pd.DataFrame, df_b: pd.DataFrame) -> list[str]:
    """
    Cerca colonne comuni che identifichino univocamente le righe in ENTRAMBI
    i DataFrame (combinazione con cardinalita' massima e duplicati minimi).
    Fallback: indice numerico (nessuna chiave).
    """
    common = [c for c in df_a.columns if c in df_b.columns]
    if not common:
        return []

    # Proviamo combinazioni dalla piu' semplice (1 col) alla piu' complessa
    from itertools import combinations

    best_key: list[str] = []
    best_score = -1.0

    for r in range(1, min(5, len(common) + 1)):
        for combo in combinations(common, r):
            combo = list(combo)
            n_unique_a = df_a[combo].drop_duplicates().shape[0]
            n_unique_b = df_b[combo].drop_duplicates().shape[0]
            # punteggio: media delle frazioni di univocita'
            score_a = n_unique_a / max(len(df_a), 1)
            score_b = n_unique_b / max(len(df_b), 1)
            score = (score_a + score_b) / 2
            if score > best_score:
                best_score = score
                best_key = combo
            if best_score >= 0.999:
                break
        if best_score >= 0.999:
            break

    # Se il punteggio e' troppo basso non ha senso usare quella chiave
    if best_score < 0.5:
        return []
    return best_key


# ---------------------------------------------------------------------------
# Confronto DataFrame
# ---------------------------------------------------------------------------

def compare_dataframes(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    join_key: list[str] | None = None,
    max_diff_rows: int = 10_000,
) -> dict:
    """
    Confronta df_a (AS-IS) con df_b (TO-BE).
    Restituisce un dizionario con:
      - summary: dict con contatori
      - cols_only_a, cols_only_b, cols_common: liste colonne
      - df_only_a, df_only_b: righe uniche
      - df_diff: righe con differenze (max_diff_rows)
      - key_used: lista colonne chiave effettivamente usata
    """
    cols_a = set(df_a.columns)
    cols_b = set(df_b.columns)
    cols_only_a = sorted(cols_a - cols_b)
    cols_only_b = sorted(cols_b - cols_a)
    cols_common = sorted(cols_a & cols_b)

    if join_key is None:
        join_key = detect_join_key(df_a[cols_common] if cols_common else df_a,
                                   df_b[cols_common] if cols_common else df_b)

    # Confronto strutturale
    struct_ok = (cols_only_a == [] and cols_only_b == [])

    if not join_key:
        # Confronto posizionale
        min_rows = min(len(df_a), len(df_b))
        common_df_a = df_a[cols_common].iloc[:min_rows].reset_index(drop=True)
        common_df_b = df_b[cols_common].iloc[:min_rows].reset_index(drop=True)
        # Confronto su valori normalizzati (elimina falsi positivi da spazi)
        norm_a = _clean_df(common_df_a)
        norm_b = _clean_df(common_df_b)
        diff_mask = (norm_a != norm_b)
        diff_rows_idx = diff_mask.any(axis=1)
        # Nel report mostra i valori originali (non normalizzati)
        df_diff_a = common_df_a[diff_rows_idx].head(max_diff_rows)
        df_diff_b = common_df_b[diff_rows_idx].head(max_diff_rows)
        df_only_a = df_a.iloc[min_rows:].head(max_diff_rows) if len(df_a) > min_rows else pd.DataFrame()
        df_only_b = df_b.iloc[min_rows:].head(max_diff_rows) if len(df_b) > min_rows else pd.DataFrame()
        n_diff = int(diff_rows_idx.sum())
        key_used = []
    else:
        # Confronto per chiave — merge sui valori normalizzati della chiave
        df_a_norm = df_a.copy()
        df_b_norm = df_b.copy()
        for k in join_key:
            if k in df_a_norm.columns:
                df_a_norm[k] = _clean_str_series(df_a_norm[k].astype(str))
            if k in df_b_norm.columns:
                df_b_norm[k] = _clean_str_series(df_b_norm[k].astype(str))

        merged = pd.merge(
            df_a_norm[cols_common + [c for c in cols_only_a]],
            df_b_norm[cols_common + [c for c in cols_only_b]],
            on=join_key, how="outer", indicator=True, suffixes=("__A", "__B"),
        )
        df_only_a = merged[merged["_merge"] == "left_only"].drop(columns=["_merge"]).head(max_diff_rows)
        df_only_b = merged[merged["_merge"] == "right_only"].drop(columns=["_merge"]).head(max_diff_rows)
        both = merged[merged["_merge"] == "both"].drop(columns=["_merge"])

        # Confronto colonne non-chiave su valori normalizzati
        compare_cols = [c for c in cols_common if c not in join_key]
        diff_rows_mask = pd.Series(False, index=both.index)
        for c in compare_cols:
            ca, cb = f"{c}__A", f"{c}__B"
            if ca in both.columns and cb in both.columns:
                val_a = _clean_str_series(both[ca].fillna("").astype(str))
                val_b = _clean_str_series(both[cb].fillna("").astype(str))
                diff_rows_mask |= (val_a != val_b)
        df_diff_a = both[diff_rows_mask][join_key + [f"{c}__A" for c in compare_cols if f"{c}__A" in both.columns]].head(max_diff_rows)
        df_diff_b = both[diff_rows_mask][join_key + [f"{c}__B" for c in compare_cols if f"{c}__B" in both.columns]].head(max_diff_rows)
        n_diff = int(diff_rows_mask.sum())
        key_used = join_key

    summary = {
        "rows_a": len(df_a),
        "rows_b": len(df_b),
        "cols_only_a": len(cols_only_a),
        "cols_only_b": len(cols_only_b),
        "cols_common": len(cols_common),
        "rows_only_a": len(df_only_a),
        "rows_only_b": len(df_only_b),
        "rows_diff": n_diff,
        "struct_ok": struct_ok,
    }
    return {
        "summary": summary,
        "cols_only_a": cols_only_a,
        "cols_only_b": cols_only_b,
        "cols_common": cols_common,
        "df_only_a": df_only_a,
        "df_only_b": df_only_b,
        "df_diff_a": df_diff_a,
        "df_diff_b": df_diff_b,
        "key_used": key_used,
    }


# ---------------------------------------------------------------------------
# Stili Excel (palette originale)
# ---------------------------------------------------------------------------
_C_HEADER   = "1F4E79"   # blu scuro
_C_SUBHDR   = "2E75B6"   # blu medio
_C_KEY_HDR  = "4472C4"   # blu chiave
_C_DIFF     = "FFE699"   # giallo differenze
_C_ONLY_AS  = "F4CCCC"   # rosa solo AS-IS
_C_ONLY_TO  = "D9EAD3"   # verde chiaro solo TO-BE
_C_OK       = "E2EFDA"   # verde ok
_C_KO       = "FCE4D6"   # arancio ko
_C_WARN     = "FFF2CC"   # giallo warn


def _xl_fill(c):
    return PatternFill("solid", fgColor=c)

def _xl_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)

def _xl_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_title(ws, text: str, n_cols: int):
    """Riga titolo blu scuro, merge su n_cols colonne."""
    ws.append([text])
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=n_cols)
    cell = ws["A1"]
    cell.fill = _xl_fill(_C_HEADER)
    cell.font = _xl_font(bold=True, color="FFFFFF", size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

def _xl_hdr_row(ws, row_idx: int, bg: str = _C_SUBHDR, fg: str = "FFFFFF"):
    """Applica stile intestazione a una riga."""
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill = _xl_fill(bg)
            cell.font = _xl_font(bold=True, color=fg)
            cell.border = _xl_border()
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center", wrap_text=True)

def _xl_data_row(ws, row_idx: int, bg: str | None = None):
    """Applica bordo (e colore opzionale) a una riga dati."""
    for cell in ws[row_idx]:
        cell.border = _xl_border()
        cell.alignment = Alignment(vertical="center")
        if bg:
            cell.fill = _xl_fill(bg)

def _xl_autofit(ws, mn: int = 8, mx: int = 50):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, mn), mx)


# ---------------------------------------------------------------------------
# Helpers analisi dati
# ---------------------------------------------------------------------------

def _infer_type(series: pd.Series) -> str:
    """Inferisce il tipo semantico di una colonna: VUOTO, INTERO, DECIMALE, DATA, TESTO."""
    s = series.dropna().head(200)
    if s.empty:
        return "VUOTO"
    num_frac = pd.to_numeric(s, errors="coerce").notna().mean()
    if num_frac > 0.9:
        return "DECIMALE" if s.str.contains(r"\.", na=False).any() else "INTERO"
    try:
        import warnings
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            pd.to_datetime(s, dayfirst=True, errors="raise")
        return "DATA"
    except Exception:
        pass
    return "TESTO"


def _col_structure(df_a: pd.DataFrame, df_b: pd.DataFrame) -> pd.DataFrame:
    """
    Restituisce un DataFrame che descrive la struttura delle colonne:
    presenza in AS-IS / TO-BE, posizione, tipo inferito, coerenza.
    """
    ca, cb = list(df_a.columns), list(df_b.columns)
    all_cols = sorted(set(ca) | set(cb),
                      key=lambda x: (ca.index(x) if x in ca else 9999))
    rows = []
    for c in all_cols:
        in_a, in_b = c in ca, c in cb
        ta = _infer_type(df_a[c]) if in_a else "—"
        tb = _infer_type(df_b[c]) if in_b else "—"
        status = "OK" if (in_a and in_b) else ("SOLO AS-IS" if in_a else "SOLO TO-BE")
        coerente = ("Si'" if (in_a and in_b and ta == tb)
                    else ("—" if not (in_a and in_b) else "No"))
        rows.append({
            "COLONNA":        c,
            "IN AS-IS":       "Si'" if in_a else "No",
            "IN TO-BE":       "Si'" if in_b else "No",
            "POS AS-IS":      ca.index(c) + 1 if in_a else "—",
            "POS TO-BE":      cb.index(c) + 1 if in_b else "—",
            "TIPO AS-IS":     ta,
            "TIPO TO-BE":     tb,
            "TIPO COERENTE":  coerente,
            "STATUS":         status,
        })
    return pd.DataFrame(rows)


def _compare_rows_wide(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    key_cols: list[str],
    value_cols: list[str],
    max_diff_rows: int,
    max_only_rows: int,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, int]:
    """
    Confronto righe, restituisce:
      both_diff : DataFrame wide (key + col__AS + col__TO) delle sole righe con diff
      only_as   : righe presenti solo in AS-IS
      only_to   : righe presenti solo in TO-BE
      n_diff_total : conteggio totale righe con differenze (non cappato)
    """
    common = key_cols + value_cols

    df_a_c = df_a[common].fillna("").reset_index(drop=True)
    df_b_c = df_b[common].fillna("").reset_index(drop=True)

    if key_cols:
        merged = pd.merge(df_a_c, df_b_c, on=key_cols,
                          how="outer", suffixes=("__AS", "__TO"),
                          indicator=True)
        as_val_cols = [c for c in merged.columns if c.endswith("__AS")]
        to_val_cols = [c for c in merged.columns if c.endswith("__TO")]
        # Teniamo SOLO le colonne del lato presente: evita duplicati e colonne
        # tutte-vuote causate dal rename di __AS e __TO sullo stesso nome base.
        only_as = (merged[merged["_merge"] == "left_only"]
                   [key_cols + as_val_cols]
                   .rename(columns=lambda c: c[:-4] if c.endswith("__AS") else c)
                   .head(max_only_rows))
        only_to = (merged[merged["_merge"] == "right_only"]
                   [key_cols + to_val_cols]
                   .rename(columns=lambda c: c[:-4] if c.endswith("__TO") else c)
                   .head(max_only_rows))
        both = merged[merged["_merge"] == "both"].drop(columns=["_merge"]).reset_index(drop=True)
    else:
        n = min(len(df_a_c), len(df_b_c))
        only_as = df_a_c.iloc[n:].head(max_only_rows).copy()
        only_to = df_b_c.iloc[n:].head(max_only_rows).copy()
        left  = df_a_c[value_cols].iloc[:n].reset_index(drop=True).add_suffix("__AS")
        right = df_b_c[value_cols].iloc[:n].reset_index(drop=True).add_suffix("__TO")
        both  = pd.concat([left, right], axis=1)

    # Maschera righe con almeno una differenza
    diff_mask = pd.Series(False, index=both.index)
    for c in value_cols:
        ca, cb = f"{c}__AS", f"{c}__TO"
        if ca in both.columns and cb in both.columns:
            diff_mask |= (both[ca].fillna("") != both[cb].fillna(""))

    n_diff_total = int(diff_mask.sum())
    both_diff = both[diff_mask].head(max_diff_rows).reset_index(drop=True)

    return both_diff, only_as, only_to, n_diff_total


def _sintesi_colonne(
    both_diff_full: pd.DataFrame,   # df wide TUTTI i match (non solo quelli con diff)
    key_cols: list[str],
    value_cols: list[str],
) -> pd.DataFrame:
    rows = []
    for c in value_cols:
        ca, cb = f"{c}__AS", f"{c}__TO"
        if ca not in both_diff_full.columns or cb not in both_diff_full.columns:
            continue
        eq   = int((both_diff_full[ca].fillna("") == both_diff_full[cb].fillna("")).sum())
        diff = len(both_diff_full) - eq
        tot  = len(both_diff_full)
        rows.append({
            "COLONNA":       c,
            "CHIAVE":        "Si'" if c in key_cols else "No",
            "RIGHE UGUALI":  eq,
            "RIGHE DIVERSE": diff,
            "TOT RIGHE":     tot,
            "% DIVERSE":     f"{diff / tot * 100:.1f}%" if tot else "0%",
            "STATO":         "OK" if diff == 0 else "DIFFERENZE",
        })
    return pd.DataFrame(rows)


def _num_diff(a, b) -> str:
    try:
        return str(float(str(a).replace(",", ".")) - float(str(b).replace(",", ".")))
    except (ValueError, TypeError):
        return ""


# ---------------------------------------------------------------------------
# Scrittura Excel — 6 fogli
# ---------------------------------------------------------------------------

def build_excel_pair(
    label: str,
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    asis_label: str,
    tobe_label: str,
    output_path: str | Path,
    join_key: list[str] | None = None,
    max_diff_rows: int = 10_000,
    max_only_rows: int = 5_000,
    log_path: str | Path | None = None,
    row_filter_desc: str | None = None,
) -> str:
    """
    Confronta df_a (AS-IS) e df_b (TO-BE) e scrive un Excel con 6 fogli:
      RIEPILOGO    — riepilogo numerico coppia
      STRUTTURA    — struttura colonne (presenza, posizione, tipo)
      SINTESI_COL  — quante righe differiscono per colonna
      DIFF         — vista wide chiave + AS-IS/TO-BE affiancati (celle diff in giallo)
      SOLO_ASIS    — righe presenti solo in AS-IS
      SOLO_TOBE    — righe presenti solo in TO-BE
    """
    output_path = Path(output_path)

    # Normalizza entrambi per sicurezza
    df_a = _clean_df(df_a)
    df_b = _clean_df(df_b)

    n_a, n_b = len(df_a), len(df_b)

    # ── Struttura colonne ────────────────────────────────────────────────────
    strut = _col_structure(df_a, df_b)
    only_a_cols = strut[strut["STATUS"] == "SOLO AS-IS"]["COLONNA"].tolist()
    only_b_cols = strut[strut["STATUS"] == "SOLO TO-BE"]["COLONNA"].tolist()

    # ── Chiave di join ────────────────────────────────────────────────────────
    if join_key is None:
        join_key = detect_join_key(df_a, df_b)
    key_cols = join_key or []
    common   = [c for c in df_a.columns if c in df_b.columns]
    value_cols = [c for c in common if c not in key_cols]

    # ── Confronto righe ────────────────────────────────────────────────────────
    both_diff, only_as, only_to, n_diff_total = _compare_rows_wide(
        df_a, df_b, key_cols, value_cols, max_diff_rows, max_only_rows
    )

    # Per SINTESI_COL servono TUTTI i match (non solo quelli con diff)
    if key_cols and not df_a.empty and not df_b.empty:
        _common_c = key_cols + value_cols
        _all_both = pd.merge(
            df_a[_common_c].fillna(""), df_b[_common_c].fillna(""),
            on=key_cols, how="inner", suffixes=("__AS", "__TO"),
        )
    else:
        _all_both = both_diff   # fallback posizionale

    sint = _sintesi_colonne(_all_both, key_cols, value_cols)
    cols_w_diff = sint[sint["STATO"] == "DIFFERENZE"]["COLONNA"].tolist() if not sint.empty else []

    ok_overall = (not only_a_cols and not only_b_cols
                  and n_diff_total == 0 and n_a == n_b)

    # ── Workbook ───────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. RIEPILOGO ──────────────────────────────────────────────────────────
    ws_r = wb.create_sheet("RIEPILOGO")
    _xl_title(ws_r, f"Riepilogo confronto AS-IS vs TO-BE  —  {label}", 11)
    headers_r = ["FILE LOGICO", "FILE AS-IS", "FILE TO-BE",
                 "RIGHE AS-IS", "RIGHE TO-BE", "DELTA RIGHE",
                 "COL SOLO AS-IS", "COL SOLO TO-BE",
                 "COLONNE CON DIFF", "RIGHE CON DIFF", "ESITO"]
    ws_r.append(headers_r)
    _xl_hdr_row(ws_r, 2)

    ws_r.append([
        label, asis_label, tobe_label,
        n_a, n_b, n_a - n_b,
        ", ".join(only_a_cols) or "—",
        ", ".join(only_b_cols) or "—",
        ", ".join(cols_w_diff) or "—",
        n_diff_total,
        "OK" if ok_overall else "DIFFERENZE",
    ])
    _xl_data_row(ws_r, 3, _C_OK if ok_overall else _C_KO)
    ws_r.append([])
    ws_r.append(["Chiave di join rilevata:",
                 ", ".join(key_cols) if key_cols else "(confronto posizionale)"])
    if row_filter_desc:
        ws_r.append(["Filtro righe applicato:", row_filter_desc])
    _xl_autofit(ws_r)
    ws_r.freeze_panes = "A3"

    # ── 2. STRUTTURA ─────────────────────────────────────────────────────────
    ws_s = wb.create_sheet("STRUTTURA")
    n_sc = len(strut.columns) if not strut.empty else 1
    _xl_title(ws_s, f"Struttura colonne  —  {label}", n_sc)
    if strut.empty:
        ws_s.append(["(nessun dato)"])
    else:
        ws_s.append(list(strut.columns))
        _xl_hdr_row(ws_s, 2)
        for _, row in strut.iterrows():
            ws_s.append(list(row))
            rn = ws_s.max_row
            st = str(row.get("STATUS", ""))
            tc = str(row.get("TIPO COERENTE", "Si'"))
            if   st == "SOLO AS-IS": _xl_data_row(ws_s, rn, _C_ONLY_AS)
            elif st == "SOLO TO-BE": _xl_data_row(ws_s, rn, _C_ONLY_TO)
            elif tc == "No":         _xl_data_row(ws_s, rn, _C_WARN)
            else:                    _xl_data_row(ws_s, rn)
    _xl_autofit(ws_s)
    ws_s.freeze_panes = "A3"

    # ── 3. SINTESI_COL ────────────────────────────────────────────────────────
    ws_sc = wb.create_sheet("SINTESI_COL")
    n_sc2 = len(sint.columns) if not sint.empty else 1
    _xl_title(ws_sc, f"Sintesi differenze per colonna  —  {label}", n_sc2)
    if sint.empty:
        ws_sc.append(["(nessun dato comparabile)"])
    else:
        ws_sc.append(list(sint.columns))
        _xl_hdr_row(ws_sc, 2)
        for _, row in sint.iterrows():
            ws_sc.append(list(row))
            rn = ws_sc.max_row
            bg = _C_KO if str(row.get("STATO", "")) == "DIFFERENZE" else _C_OK
            _xl_data_row(ws_sc, rn, bg)
    _xl_autofit(ws_sc)
    ws_sc.freeze_panes = "A3"

    # ── 4. DIFF ────────────────────────────────────────────────────────────────
    ws_d = wb.create_sheet("DIFF")
    if both_diff.empty or not value_cols:
        _xl_title(ws_d, f"Differenze riga x colonna  —  {label}", 1)
        ws_d.append(["(nessuna differenza rilevata)"])
    else:
        truncated = n_diff_total > max_diff_rows
        note = (f" — prime {max_diff_rows:,} su {n_diff_total:,} righe con diff"
                if truncated else f" — {len(both_diff):,} righe con diff")
        # Costruisci intestazione wide: chiave | col [AS-IS] | col [TO-BE] | DIFF col | ...
        headers_d = list(key_cols)
        for c in value_cols:
            headers_d += [f"{c} [AS-IS]", f"{c} [TO-BE]", f"DIFF {c}"]

        _xl_title(ws_d, f"Differenze riga x colonna  —  {label}{note}", len(headers_d))
        ws_d.append(headers_d)
        _xl_hdr_row(ws_d, 2)
        # Intestazioni chiave con colore distinto
        for i, _ in enumerate(key_cols, 1):
            ws_d.cell(row=2, column=i).fill = _xl_fill(_C_KEY_HDR)

        for _, row in both_diff.iterrows():
            row_vals = [row.get(k, "") for k in key_cols]
            for c in value_cols:
                va = row.get(f"{c}__AS", "")
                vb = row.get(f"{c}__TO", "")
                row_vals += [va, vb, _num_diff(va, vb)]
            ws_d.append(row_vals)
            rn = ws_d.max_row
            _xl_data_row(ws_d, rn)
            # Evidenzia in giallo le celle AS-IS e TO-BE dove c'e' differenza
            col_offset = len(key_cols) + 1
            for c in value_cols:
                va = str(row.get(f"{c}__AS", "")).strip()
                vb = str(row.get(f"{c}__TO", "")).strip()
                if va != vb:
                    ws_d.cell(rn, col_offset).fill     = _xl_fill(_C_DIFF)
                    ws_d.cell(rn, col_offset + 1).fill = _xl_fill(_C_DIFF)
                col_offset += 3

        _xl_autofit(ws_d)
        if key_cols:
            ws_d.freeze_panes = f"{get_column_letter(len(key_cols) + 1)}3"

    # ── 5. SOLO_ASIS ──────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("SOLO_ASIS")
    nc_a = max(len(only_as.columns), 1) if not only_as.empty else 1
    trunc_a = len(only_as) >= max_only_rows
    title_a = (f"Solo in AS-IS  —  {label}"
               + (f"  (prime {max_only_rows:,} righe)" if trunc_a else ""))
    _xl_title(ws_a, title_a, nc_a)
    if only_as.empty:
        ws_a.append(["(nessuna riga esclusiva in AS-IS)"])
    else:
        ws_a.append(list(only_as.columns))
        _xl_hdr_row(ws_a, 2)
        for _, row in only_as.iterrows():
            ws_a.append(list(row))
            _xl_data_row(ws_a, ws_a.max_row, _C_ONLY_AS)
    _xl_autofit(ws_a)

    # ── 6. SOLO_TOBE ──────────────────────────────────────────────────────────
    ws_t = wb.create_sheet("SOLO_TOBE")
    nc_t = max(len(only_to.columns), 1) if not only_to.empty else 1
    trunc_t = len(only_to) >= max_only_rows
    title_t = (f"Solo in TO-BE  —  {label}"
               + (f"  (prime {max_only_rows:,} righe)" if trunc_t else ""))
    _xl_title(ws_t, title_t, nc_t)
    if only_to.empty:
        ws_t.append(["(nessuna riga esclusiva in TO-BE)"])
    else:
        ws_t.append(list(only_to.columns))
        _xl_hdr_row(ws_t, 2)
        for _, row in only_to.iterrows():
            ws_t.append(list(row))
            _xl_data_row(ws_t, ws_t.max_row, _C_ONLY_TO)
    _xl_autofit(ws_t)

    wb.save(output_path)
    return str(output_path)


# ---------------------------------------------------------------------------
# Rilevamento righe malformate (saltate da on_bad_lines='skip')
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Supporto file grandi — streaming + SQLite temporaneo
# ---------------------------------------------------------------------------

_LARGE_FILE_THRESHOLD = 200 * 1024 * 1024   # 200 MB → modalità streaming


def _source_size(path_str: str) -> int:
    """Dimensione stimata in byte del CSV (decompressa se inside ZIP)."""
    try:
        if "::" in path_str:
            zp, ze = path_str.split("::", 1)
            with zipfile.ZipFile(zp) as zf:
                info = zf.getinfo(ze)
                return info.file_size or info.compress_size
        else:
            return Path(path_str).stat().st_size
    except Exception:
        return 0


def _iter_csv_chunks(
    path_str: str,
    sep: str,
    chunk_size: int = 50_000,
):
    """
    Genera coppie (col_names, chunk_df) leggendo un CSV a blocchi senza
    caricarlo interamente in RAM.  Funziona sia su file CSV che su ZIP entry.
    """
    # ── Rileva encoding e header dalla prima lettura ──────────────────────────
    if "::" in path_str:
        zp, ze = path_str.split("::", 1)
        with zipfile.ZipFile(zp) as zf:
            with zf.open(ze) as fh:
                first_bytes = fh.read(65536)
    else:
        with open(path_str, "rb") as fh:
            first_bytes = fh.read(65536)

    enc        = _detect_encoding(first_bytes)
    first_text = first_bytes.decode(enc, errors="replace")
    has_hdr    = _has_header_from_text(first_text, sep)
    header_arg = 0 if has_hdr else None
    eng        = "python" if len(sep) > 1 else "c"
    sep_param  = re.escape(sep) if eng == "python" else sep

    # ── Apre lo stream per la lettura completa ────────────────────────────────
    if "::" in path_str:
        zp, ze = path_str.split("::", 1)
        outer_zf = zipfile.ZipFile(zp)
        raw_fh   = outer_zf.open(ze)
    else:
        outer_zf = None
        raw_fh   = open(path_str, "rb")

    try:
        import io as _io
        text_fh = _io.TextIOWrapper(raw_fh, encoding=enc, errors="replace")
        try:
            reader = pd.read_csv(
                text_fh,
                sep=sep_param,
                header=header_arg,
                dtype=str,
                chunksize=chunk_size,
                engine=eng,
                on_bad_lines="skip",
                keep_default_na=False,
                skipinitialspace=True,
                index_col=False,   # evita shift quando i dati hanno trailing sep
            )
            col_names: list[str] = []
            for i, chunk in enumerate(reader):
                if i == 0:
                    if header_arg is None:
                        chunk.columns = [f"Col_{j+1}" for j in range(len(chunk.columns))]
                    chunk.columns = _dedup_columns([str(c).strip() for c in chunk.columns])
                    chunk = _strip_sep_prefixes(chunk, sep)
                    chunk = chunk[[c for c in chunk.columns if not _is_artifact_col(c)]]
                    col_names = list(chunk.columns)
                else:
                    # Allinea colonne ai nomi rilevati nel primo chunk
                    if len(chunk.columns) >= len(col_names):
                        chunk = chunk.iloc[:, :len(col_names)]
                    chunk.columns = col_names[:len(chunk.columns)]
                chunk = _clean_df(chunk)
                yield col_names, chunk
        finally:
            try:
                text_fh.detach()   # evita doppia chiusura di raw_fh
            except Exception:
                pass
    finally:
        try:
            raw_fh.close()
        except Exception:
            pass
        if outer_zf:
            try:
                outer_zf.close()
            except Exception:
                pass


def _load_to_duckdb(
    path_str: str,
    sep: str,
    db_path: str,
    table: str,
    chunk_size: int = 100_000,
    log_cb: Callable[[str], None] | None = None,
    row_filter: "dict[str, str] | None" = None,
) -> tuple[list[str], int]:
    """
    Carica un CSV (o ZIP entry) in una tabella DuckDB su file, a blocchi.
    Ogni chiamata usa una connessione DuckDB dedicata (thread-safe).
    Aggiunge la colonna sentinella _fc_exists=1 per rilevare LEFT JOIN miss.
    Restituisce (nomi_colonne_normalizzati, n_righe_totali).
    """
    import duckdb as _duckdb
    conn   = _duckdb.connect(db_path)
    total  = 0
    col_names: list[str] = []
    first  = True

    try:
        for cols, chunk in _iter_csv_chunks(path_str, sep, chunk_size):
            if not col_names:
                col_names = cols
            if row_filter:
                for _fc, _fv in row_filter.items():
                    if _fc in chunk.columns:
                        chunk = chunk[chunk[_fc].str.strip() == _fv.strip()]
                if chunk.empty:
                    continue
            chunk["_fc_exists"] = 1   # sentinella per LEFT JOIN IS NULL
            if first:
                conn.execute(f'CREATE TABLE "{table}" AS SELECT * FROM chunk')
                first = False
            else:
                conn.execute(f'INSERT INTO "{table}" SELECT * FROM chunk')
            total += len(chunk)
            if log_cb and total % (chunk_size * 4) < chunk_size:
                log_cb(f"    {table.upper()}: {total:,} righe caricate...")
    finally:
        conn.close()

    if log_cb:
        log_cb(f"    {table.upper()}: completato — {total:,} righe, {len(col_names)} colonne")
    return col_names, total


def _q(name: str) -> str:
    """Quota un identificatore SQLite."""
    return '"' + name.replace('"', '""') + '"'


def build_excel_pair_large(
    label: str,
    path_a: str,
    path_b: str,
    sep_a: str,
    sep_b: str,
    asis_label: str,
    tobe_label: str,
    output_path: "str | Path",
    join_key: "list[str] | None" = None,
    max_diff_rows: int = 10_000,
    max_only_rows: int = 5_000,
    chunk_size: int = 100_000,
    log_cb: "Callable[[str], None] | None" = None,
    row_filter: "dict[str, str] | None" = None,
) -> str:
    """
    Confronta due grandi CSV via DuckDB (OLAP, colonnare, multithreaded).
    Carica AS-IS e TO-BE in PARALLELO in due file DuckDB separati,
    poi esegue le query analitiche con hash join vettorizzato.
    Produce il medesimo Excel a 6 fogli di build_excel_pair.
    """
    import duckdb as _duckdb
    import tempfile as _tempfile
    import threading as _threading

    def _log(msg: str):
        if log_cb:
            log_cb(msg)

    output_path = Path(output_path)

    # ── File DuckDB temporanei (uno per lato, per caricamento parallelo) ───────
    fd_a, db_path_a = _tempfile.mkstemp(suffix=".duckdb", prefix="flowcheck_a_")
    fd_b, db_path_b = _tempfile.mkstemp(suffix=".duckdb", prefix="flowcheck_b_")
    # mkstemp lascia il fd aperto — chiudiamolo prima di unlink (richiesto su Windows)
    os.close(fd_a); os.close(fd_b)
    os.unlink(db_path_a)
    os.unlink(db_path_b)

    load_results: dict = {}
    load_errors:  dict = {}

    def _load_side(side: str, path: str, sep: str, db_path: str, table: str, label_str: str):
        try:
            cols, n = _load_to_duckdb(path, sep, db_path, table, chunk_size, _log, row_filter)
            load_results[side] = (cols, n)
        except Exception as exc:
            load_errors[side] = exc

    # ── Caricamento PARALLELO ─────────────────────────────────────────────────
    _log("    Caricamento AS-IS e TO-BE in parallelo (DuckDB)...")
    t_a = _threading.Thread(target=_load_side,
                            args=("a", path_a, sep_a, db_path_a, "asis", "AS-IS"))
    t_b = _threading.Thread(target=_load_side,
                            args=("b", path_b, sep_b, db_path_b, "tobe", "TO-BE"))
    t_a.start(); t_b.start()
    t_a.join();  t_b.join()

    if load_errors:
        raise RuntimeError(f"Errore caricamento: {load_errors}")

    cols_a, n_a = load_results["a"]
    cols_b, n_b = load_results["b"]

    try:
        # ── Connessione principale: ATTACH di entrambi i DB ───────────────────
        conn = _duckdb.connect()
        conn.execute(f"ATTACH '{db_path_a}' AS dba (READ_ONLY)")
        conn.execute(f"ATTACH '{db_path_b}' AS dbb (READ_ONLY)")
        conn.execute("CREATE VIEW asis AS SELECT * FROM dba.asis")
        conn.execute("CREATE VIEW tobe AS SELECT * FROM dbb.tobe")

        cols_a_set  = set(cols_a)
        cols_b_set  = set(cols_b)
        common      = [c for c in cols_a if c in cols_b_set]
        only_a_cols = sorted(cols_a_set - cols_b_set)
        only_b_cols = sorted(cols_b_set - cols_a_set)

        # Colonne utente (escludi sentinella interna)
        user_cols_a = [c for c in cols_a if c != "_fc_exists"]
        user_cols_b = [c for c in cols_b if c != "_fc_exists"]
        common_user = [c for c in user_cols_a if c in set(user_cols_b)]

        # ── Auto-detect chiave su campione ────────────────────────────────────
        if join_key is None:
            if common_user:
                sel_c = ", ".join(_q(c) for c in common_user)
                smp_a = conn.execute(f"SELECT {sel_c} FROM asis LIMIT 100000").df()
                smp_b = conn.execute(f"SELECT {sel_c} FROM tobe LIMIT 100000").df()
                join_key = detect_join_key(smp_a, smp_b)
                del smp_a, smp_b
            else:
                join_key = []

        key_cols   = join_key or []
        value_cols = [c for c in common_user if c not in key_cols]
        _log(f"    Chiave: {', '.join(key_cols) if key_cols else '(nessuna — confronto posizionale)'}")
        _log("    Avvio query DuckDB (hash join vettorizzato)...")

        # ── Helper SQL (DuckDB: COALESCE ok, hash join non usa indici) ────────
        def kj(ta: str = "a", tb: str = "b") -> str:
            return " AND ".join(
                f"COALESCE({ta}.{_q(k)},'') = COALESCE({tb}.{_q(k)},'')"
                for k in key_cols
            )

        sint_rows: list[dict] = []
        n_both = n_only_a = n_only_b = n_diff = 0

        if key_cols:
            # ── SINTESI completa in UNA SOLA query DuckDB ─────────────────────
            # DuckDB: vectorized, columnar, multithreaded → milliseconds su <10M righe
            # COUNT(*) FILTER è sintassi DuckDB nativa (non disponibile in SQLite)
            cases = ",\n    ".join(
                f"SUM(CASE WHEN b._fc_exists IS NOT NULL AND "
                f"COALESCE(a.{_q(c)},'') != COALESCE(b.{_q(c)},'') "
                f"THEN 1 ELSE 0 END) AS {_q('_d_' + c)}"
                for c in value_cols
            )
            select_clause = (
                "COUNT(*) FILTER (WHERE b._fc_exists IS NOT NULL) AS _n_both,\n"
                "    COUNT(*) FILTER (WHERE b._fc_exists IS NULL)     AS _n_only_a"
                + (f",\n    {cases}" if cases else "")
            )
            _log("    SINTESI (query unica su tutte le colonne)...")
            row = conn.execute(
                f"SELECT {select_clause}\n"
                f"FROM asis a LEFT JOIN tobe b ON {kj()}"
            ).fetchone()
            n_both   = row[0] or 0
            n_only_a = row[1] or 0
            for j, c in enumerate(value_cols):
                nd = row[2 + j] or 0
                sint_rows.append({
                    "COLONNA": c, "CHIAVE": "No",
                    "RIGHE UGUALI": n_both - nd, "RIGHE DIVERSE": nd,
                    "TOT": n_both,
                    "STATO": "OK" if nd == 0 else "DIFFERENZE",
                })

            # ── n_only_b ─────────────────────────────────────────────────────
            _log("    Calcolo righe solo TO-BE...")
            n_only_b = conn.execute(
                f"SELECT COUNT(*) FILTER (WHERE a._fc_exists IS NULL)\n"
                f"FROM tobe b LEFT JOIN asis a ON {kj('b', 'a')}"
            ).fetchone()[0]

            # ── n_diff: sulle sole colonne con differenze ─────────────────────
            diff_cols = [sr["COLONNA"] for sr in sint_rows if sr["STATO"] == "DIFFERENZE"]
            if diff_cols:
                diff_cond = " OR ".join(
                    f"COALESCE(a.{_q(c)},'') != COALESCE(b.{_q(c)},'')"
                    for c in diff_cols[:60]
                )
                n_diff = conn.execute(
                    f"SELECT COUNT(*) FROM asis a\n"
                    f"INNER JOIN tobe b ON {kj()} WHERE {diff_cond}"
                ).fetchone()[0]
        else:
            n_only_a = max(0, n_a - n_b)
            n_only_b = max(0, n_b - n_a)

        ok_overall = (n_only_a == 0 and n_only_b == 0 and n_diff == 0 and n_a == n_b)

        zero_match_warning = (key_cols and n_both == 0 and n_a > 0 and n_b > 0)
        if zero_match_warning:
            _log("    [ATTENZIONE] La chiave di join non ha prodotto nessun match "
                 f"({', '.join(key_cols)}). Verificare che i valori chiave coincidano "
                 "tra AS-IS e TO-BE (case, spazi, prefissi £/#).")

        # ── Workbook ──────────────────────────────────────────────────────────
        wb  = openpyxl.Workbook()
        wb.remove(wb.active)

        # 1. RIEPILOGO
        ws_r = wb.create_sheet("RIEPILOGO")
        _xl_title(ws_r, f"Riepilogo confronto AS-IS vs TO-BE  —  {label}", 11)
        ws_r.append(["FILE LOGICO", "FILE AS-IS", "FILE TO-BE",
                     "RIGHE AS-IS", "RIGHE TO-BE", "DELTA RIGHE",
                     "COL SOLO AS-IS", "COL SOLO TO-BE",
                     "COLONNE CON DIFF", "RIGHE CON DIFF", "ESITO"])
        _xl_hdr_row(ws_r, 2)
        cols_w_diff = [r["COLONNA"] for r in sint_rows if r["STATO"] == "DIFFERENZE"]
        ws_r.append([
            label, asis_label, tobe_label,
            n_a, n_b, n_a - n_b,
            ", ".join(only_a_cols) or "—",
            ", ".join(only_b_cols) or "—",
            ", ".join(cols_w_diff[:10]) + ("…" if len(cols_w_diff) > 10 else "") or "—",
            n_diff,
            "OK" if ok_overall else "DIFFERENZE",
        ])
        _xl_data_row(ws_r, 3, _C_OK if ok_overall else _C_KO)
        ws_r.append([])
        ws_r.append(["Chiave di join:", ", ".join(key_cols) if key_cols else "(nessuna — confronto posizionale)"])
        ws_r.append(["Righe con match (join):", n_both if key_cols else "n/a"])
        ws_r.append(["Modalità elaborazione:", "STREAM  (file grande — confronto via DuckDB)"])
        if row_filter:
            _filter_desc = "  AND  ".join(f"{k} = '{v}'" for k, v in row_filter.items())
            ws_r.append(["Filtro righe applicato:", _filter_desc])
        if zero_match_warning:
            ws_r.append([])
            warn_row = ["ATTENZIONE: la chiave di join non ha trovato nessun match tra AS-IS e TO-BE. "
                        f"Chiave usata: {', '.join(key_cols)}. "
                        "Possibili cause: valori con prefissi diversi, case sensitivity, "
                        "separatore rilevato in modo errato."]
            ws_r.append(warn_row)
            warn_rn = ws_r.max_row
            ws_r.cell(warn_rn, 1).fill = _xl_fill(_C_KO)
            ws_r.cell(warn_rn, 1).font = Font(color=_CLR["err_fg"], bold=True)
        _xl_autofit(ws_r)
        ws_r.freeze_panes = "A3"

        # 2. STRUTTURA
        ws_s = wb.create_sheet("STRUTTURA")
        _xl_title(ws_s, f"Struttura colonne  —  {label}", 6)
        ws_s.append(["COLONNA", "IN AS-IS", "IN TO-BE", "POS AS-IS", "POS TO-BE", "STATUS"])
        _xl_hdr_row(ws_s, 2)
        all_cols = user_cols_a + [c for c in user_cols_b if c not in set(user_cols_a)]
        for c in all_cols:
            ia = c in set(user_cols_a)
            ib = c in set(user_cols_b)
            st = "OK" if (ia and ib) else ("SOLO AS-IS" if ia else "SOLO TO-BE")
            ws_s.append([c,
                         "Si" if ia else "No", "Si" if ib else "No",
                         user_cols_a.index(c) + 1 if ia else "—",
                         user_cols_b.index(c) + 1 if ib else "—",
                         st])
            rn = ws_s.max_row
            if   st == "SOLO AS-IS": _xl_data_row(ws_s, rn, _C_ONLY_AS)
            elif st == "SOLO TO-BE": _xl_data_row(ws_s, rn, _C_ONLY_TO)
            else:                    _xl_data_row(ws_s, rn)
        _xl_autofit(ws_s)
        ws_s.freeze_panes = "A3"

        # 3. SINTESI_COL
        ws_sc = wb.create_sheet("SINTESI_COL")
        _xl_title(ws_sc, f"Sintesi differenze per colonna  —  {label}", 6)
        ws_sc.append(["COLONNA", "CHIAVE", "RIGHE UGUALI", "RIGHE DIVERSE", "TOT RIGHE", "STATO"])
        _xl_hdr_row(ws_sc, 2)
        if sint_rows:
            for sr in sint_rows:
                bg = _C_OK if sr["STATO"] == "OK" else _C_KO
                ws_sc.append([sr["COLONNA"], sr["CHIAVE"],
                               sr["RIGHE UGUALI"], sr["RIGHE DIVERSE"], sr["TOT"], sr["STATO"]])
                _xl_data_row(ws_sc, ws_sc.max_row, bg)
        elif zero_match_warning:
            msg = (f"Statistiche non disponibili: chiave ({', '.join(key_cols)}) "
                   "non ha trovato nessun match.")
            ws_sc.append([msg])
            ws_sc.cell(ws_sc.max_row, 1).fill = _xl_fill(_C_KO)
            ws_sc.cell(ws_sc.max_row, 1).font = Font(color=_CLR["err_fg"], bold=True)
        else:
            ws_sc.append(["(dati non disponibili senza chiave di join)"])
        _xl_autofit(ws_sc)
        ws_sc.freeze_panes = "A3"

        # 4. DIFF
        ws_d = wb.create_sheet("DIFF")
        if key_cols and value_cols and n_diff > 0:
            diff_cond = " OR ".join(
                f"COALESCE(a.{_q(c)},'') != COALESCE(b.{_q(c)},'')"
                for c in (diff_cols if diff_cols else value_cols)[:60]
            )
            sel_k  = ", ".join(f"a.{_q(k)}" for k in key_cols)
            sel_av = ", ".join(f"a.{_q(c)} AS {_q(c + '__AS')}" for c in value_cols)
            sel_bv = ", ".join(f"b.{_q(c)} AS {_q(c + '__TO')}" for c in value_cols)
            diff_df = conn.execute(
                f"SELECT {sel_k}, {sel_av}, {sel_bv}\n"
                f"FROM asis a INNER JOIN tobe b ON {kj()}\n"
                f"WHERE {diff_cond}\nLIMIT {max_diff_rows}"
            ).df()
            note = (f" — prime {max_diff_rows:,} su {n_diff:,}" if n_diff > max_diff_rows
                    else f" — {n_diff:,} righe con diff")
            hdr_d = list(key_cols)
            for c in value_cols:
                hdr_d += [f"{c} [AS-IS]", f"{c} [TO-BE]", f"DIFF {c}"]
            _xl_title(ws_d, f"Differenze  —  {label}{note}", len(hdr_d))
            ws_d.append(hdr_d)
            _xl_hdr_row(ws_d, 2)
            for _, row in diff_df.iterrows():
                rv = [row.get(k, "") for k in key_cols]
                for c in value_cols:
                    va = row.get(f"{c}__AS", "")
                    vb = row.get(f"{c}__TO", "")
                    rv += [va, vb, _num_diff(va, vb)]
                ws_d.append(rv)
                rn = ws_d.max_row
                _xl_data_row(ws_d, rn)
                off = len(key_cols) + 1
                for c in value_cols:
                    if str(row.get(f"{c}__AS", "")).strip() != str(row.get(f"{c}__TO", "")).strip():
                        ws_d.cell(rn, off).fill     = _xl_fill(_C_DIFF)
                        ws_d.cell(rn, off + 1).fill = _xl_fill(_C_DIFF)
                    off += 3
            _xl_autofit(ws_d)
            if key_cols:
                ws_d.freeze_panes = f"{get_column_letter(len(key_cols)+1)}3"
        else:
            _xl_title(ws_d, f"Differenze  —  {label}", 1)
            if zero_match_warning:
                msg = ("Impossibile mostrare differenze: chiave di join senza match. "
                       "Verificare prefissi o separatore.")
            elif n_diff == 0 and key_cols:
                msg = "(nessuna differenza rilevata — tutti i match concordano)"
            else:
                msg = "(confronto non disponibile senza chiave di join)"
            ws_d.append([msg])
            if zero_match_warning:
                ws_d.cell(ws_d.max_row, 1).fill = _xl_fill(_C_KO)
                ws_d.cell(ws_d.max_row, 1).font = Font(color=_CLR["err_fg"], bold=True)

        # 5. SOLO_ASIS
        ws_a = wb.create_sheet("SOLO_ASIS")
        if key_cols:
            sel = ", ".join(f"a.{_q(c)}" for c in user_cols_a)
            only_a_df = conn.execute(
                f"SELECT {sel} FROM asis a\n"
                f"LEFT JOIN tobe b ON {kj()}\n"
                f"WHERE b._fc_exists IS NULL LIMIT {max_only_rows}"
            ).df()
        else:
            only_a_df = conn.execute(
                f"SELECT {', '.join(_q(c) for c in user_cols_a)} FROM asis LIMIT {max_only_rows}"
            ).df()
        trunc_a = n_only_a > max_only_rows
        ttl_a = (f"Solo in AS-IS  —  {label}"
                 + (f"  (prime {max_only_rows:,} su {n_only_a:,})" if trunc_a else ""))
        _xl_title(ws_a, ttl_a, max(len(only_a_df.columns), 1))
        if only_a_df.empty:
            ws_a.append(["(nessuna riga esclusiva in AS-IS)"])
        else:
            ws_a.append(list(only_a_df.columns))
            _xl_hdr_row(ws_a, 2)
            for _, row in only_a_df.iterrows():
                ws_a.append(list(row))
                _xl_data_row(ws_a, ws_a.max_row, _C_ONLY_AS)
        _xl_autofit(ws_a)

        # 6. SOLO_TOBE
        ws_t = wb.create_sheet("SOLO_TOBE")
        if key_cols:
            sel = ", ".join(f"b.{_q(c)}" for c in user_cols_b)
            only_b_df = conn.execute(
                f"SELECT {sel} FROM tobe b\n"
                f"LEFT JOIN asis a ON {kj('b', 'a')}\n"
                f"WHERE a._fc_exists IS NULL LIMIT {max_only_rows}"
            ).df()
        else:
            only_b_df = conn.execute(
                f"SELECT {', '.join(_q(c) for c in user_cols_b)} FROM tobe LIMIT {max_only_rows}"
            ).df()
        trunc_b = n_only_b > max_only_rows
        ttl_b = (f"Solo in TO-BE  —  {label}"
                 + (f"  (prime {max_only_rows:,} su {n_only_b:,})" if trunc_b else ""))
        _xl_title(ws_t, ttl_b, max(len(only_b_df.columns), 1))
        if only_b_df.empty:
            ws_t.append(["(nessuna riga esclusiva in TO-BE)"])
        else:
            ws_t.append(list(only_b_df.columns))
            _xl_hdr_row(ws_t, 2)
            for _, row in only_b_df.iterrows():
                ws_t.append(list(row))
                _xl_data_row(ws_t, ws_t.max_row, _C_ONLY_TO)
        _xl_autofit(ws_t)

        conn.close()
        wb.save(output_path)
        return str(output_path)

    finally:
        for p in (db_path_a, db_path_b):
            for ext in ("", ".wal"):
                try:
                    os.unlink(p + ext)
                except Exception:
                    pass


def _get_raw_content(path_str: str, zip_entry: str | None = None) -> str:
    """Legge il contenuto grezzo da file CSV o da una entry ZIP, con encoding auto-detect."""
    if zip_entry:
        with zipfile.ZipFile(path_str) as zf:
            with zf.open(zip_entry) as fh:
                raw_bytes = fh.read()
        enc = _detect_encoding(raw_bytes)
        return raw_bytes.decode(enc, errors="replace")
    else:
        with open(path_str, "rb") as fh:
            raw_bytes = fh.read()
        enc = _detect_encoding(raw_bytes)
        return raw_bytes.decode(enc, errors="replace")


def _find_bad_lines(raw: str, sep: str) -> list[tuple[int, str]]:
    """
    Rileva le righe che pandas skipperebbe con on_bad_lines='skip',
    usando pandas stesso con on_bad_lines='warn' per catturare i ParserWarning.
    Questo approccio e' immune ai trailing-separator e ad altre varianti
    perche' e' pandas stesso a decidere quali righe sono malformate.

    Restituisce [(numero_riga_1based, contenuto_riga)].
    """
    import io as _io
    import re as _re
    import warnings as _warnings

    raw_lines = raw.splitlines()
    bad: list[tuple[int, str]] = []

    buf = _io.StringIO(raw)
    esc = _re.escape(sep) if len(sep) > 1 else sep

    with _warnings.catch_warnings(record=True) as caught:
        _warnings.simplefilter("always")
        try:
            pd.read_csv(
                buf,
                sep=esc,
                engine="python",          # python engine emette ParserWarning
                on_bad_lines="warn",
                dtype=str,
                keep_default_na=False,
            )
        except Exception:
            pass

    for w in caught:
        if not issubclass(w.category, (pd.errors.ParserWarning, UserWarning)):
            continue
        msg = str(w.message)
        # Formato python engine: "Skipping line N: expected X fields, saw Y"
        m = (_re.search(r"[Ss]kipping\s+(?:row|line)\s+(\d+)", msg) or
             _re.search(r"in line\s+(\d+)", msg))
        if m:
            ln = int(m.group(1))          # numero riga 1-based nel file
            if 1 <= ln <= len(raw_lines):
                bad.append((ln, raw_lines[ln - 1]))

    return bad


# ---------------------------------------------------------------------------
# Issue log — raccolta dati e generazione Excel riepilogo anomalie
# ---------------------------------------------------------------------------

def _detect_whitespace_in_raw(filepath: str | Path,
                               zip_entry: str | None,
                               sep: str,
                               n_lines: int = 100) -> int:
    """
    Legge le prime n_lines righe raw (senza pulizia) e conta i campi
    che hanno spazi leading/trailing o caratteri non-breaking (\xa0).
    """
    fp = Path(filepath)
    lines = _read_first_lines(fp, n=n_lines + 1, zip_entry=zip_entry)
    if len(lines) > 1:
        lines = lines[1:]          # salta intestazione
    count = 0
    for line in lines:
        for field in line.rstrip("\n\r").split(sep):
            raw = field.strip('"').strip("'")
            if raw != raw.strip() or "\xa0" in raw:
                count += 1
    return count


def _collect_pair_issues(
    label: str,
    asis_label: str,
    tobe_label: str,
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    sep_a: str,
    sep_b: str,
    n_ws_a: int,
    n_ws_b: int,
    bad_lines_a: list[tuple[int, str]] | None = None,
    bad_lines_b: list[tuple[int, str]] | None = None,
) -> dict:
    """Raccoglie tutte le anomalie di una coppia AS-IS / TO-BE."""
    strut = _col_structure(df_a, df_b)
    only_a_cols = strut[strut["STATUS"] == "SOLO AS-IS"]["COLONNA"].tolist()
    only_b_cols = strut[strut["STATUS"] == "SOLO TO-BE"]["COLONNA"].tolist()
    type_mm = (strut[(strut["STATUS"] == "OK") & (strut["TIPO COERENTE"] == "No")]
               [["COLONNA", "TIPO AS-IS", "TIPO TO-BE"]]
               .to_dict("records"))

    key_cols   = detect_join_key(df_a, df_b)
    common     = [c for c in df_a.columns if c in df_b.columns]
    value_cols = [c for c in common if c not in key_cols]

    # Conta solo (max 1 riga, vogliamo solo i contatori)
    _, only_as_df, only_to_df, n_diff = _compare_rows_wide(
        df_a, df_b, key_cols, value_cols,
        max_diff_rows=1, max_only_rows=5_000,
    )

    return {
        "label":           label,
        "asis_label":      asis_label,
        "tobe_label":      tobe_label,
        "sep_asis":        sep_a,
        "sep_tobe":        sep_b,
        "sep_ok":          sep_a == sep_b,
        "n_a":             len(df_a),
        "n_b":             len(df_b),
        "cols_only_a":     only_a_cols,
        "cols_only_b":     only_b_cols,
        "type_mismatches": type_mm,
        "n_diff_rows":     n_diff,
        "n_only_a":        len(only_as_df),
        "n_only_b":        len(only_to_df),
        "n_ws_a":          n_ws_a,
        "n_ws_b":          n_ws_b,
        "bad_lines_a":     bad_lines_a or [],
        "bad_lines_b":     bad_lines_b or [],
        "error":           None,
    }


def build_issue_log(issue_records: list[dict], output_path: str | Path) -> str:
    """
    Genera un Excel riepilogo anomalie — da allegare alle mail.

    Fogli:
      RIEPILOGO          — una riga per coppia con tutti i contatori
      SEPARATORI         — solo coppie con separatore AS-IS != TO-BE
      ANOMALIE_STRUTTURA — colonne mancanti o in piu' rispetto all'altra versione
      CONFORMITA_TIPI    — colonne con tipo inferito diverso tra AS-IS e TO-BE
    """
    output_path = Path(output_path)
    ts_label = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # helper: riga con sfondo condizionale
    def _row_colored(ws, values: list, bg: str | None):
        ws.append(values)
        _xl_data_row(ws, ws.max_row, bg)

    # ── 1. RIEPILOGO ──────────────────────────────────────────────────────────
    ws_r = wb.create_sheet("RIEPILOGO")
    n_pairs = len(issue_records)
    _xl_title(ws_r, f"Issue Log — Confronto AS-IS vs TO-BE  ({ts_label})", 14)
    hdr = ["FILE LOGICO", "FILE AS-IS", "FILE TO-BE",
           "SEP AS-IS", "SEP TO-BE", "SEP DIVERSO?",
           "RIGHE AS-IS", "RIGHE TO-BE", "DELTA RIGHE",
           "COL MANCANTI", "RIGHE CON DIFF",
           "RIGHE SOLO AS-IS", "RIGHE SOLO TO-BE",
           "ESITO"]
    ws_r.append(hdr)
    _xl_hdr_row(ws_r, 2)

    for rec in issue_records:
        if rec.get("error"):
            _row_colored(ws_r, [
                rec["label"], rec.get("asis_label",""), rec.get("tobe_label",""),
                "—","—","—","—","—","—","—","—","—","—",
                f"ERRORE: {rec['error']}",
            ], _C_WARN)
            continue

        n_col_miss = len(rec["cols_only_a"]) + len(rec["cols_only_b"])
        sep_div    = "SI'" if not rec["sep_ok"] else "No"
        delta      = rec["n_a"] - rec["n_b"]
        has_issue  = (not rec["sep_ok"] or n_col_miss > 0
                      or rec["n_diff_rows"] > 0
                      or rec["n_only_a"] > 0 or rec["n_only_b"] > 0)
        esito      = "OK" if not has_issue else "ANOMALIE"

        _row_colored(ws_r, [
            rec["label"], rec["asis_label"], rec["tobe_label"],
            rec["sep_asis"], rec["sep_tobe"], sep_div,
            rec["n_a"], rec["n_b"], delta,
            n_col_miss,
            rec["n_diff_rows"],
            rec["n_only_a"], rec["n_only_b"],
            esito,
        ], _C_OK if esito == "OK" else _C_KO)

    # Riga totali
    ws_r.append([])
    ws_r.append(["Totale coppie analizzate:", n_pairs])
    ws_r.append(["Coppie con anomalie:",
                 sum(1 for r in issue_records
                     if not r.get("error") and (
                         not r["sep_ok"]
                         or len(r["cols_only_a"]) + len(r["cols_only_b"]) > 0
                         or r["n_diff_rows"] > 0
                         or r["n_only_a"] > 0 or r["n_only_b"] > 0
                     ))])
    _xl_autofit(ws_r)
    ws_r.freeze_panes = "A3"

    # ── 2. SEPARATORI ─────────────────────────────────────────────────────────
    ws_sep = wb.create_sheet("SEPARATORI")
    _xl_title(ws_sep, "File con separatore CSV diverso tra AS-IS e TO-BE", 6)
    ws_sep.append(["FILE LOGICO", "FILE AS-IS", "FILE TO-BE",
                   "SEP AS-IS", "SEP TO-BE", "NOTA"])
    _xl_hdr_row(ws_sep, 2)

    sep_anomalies = [r for r in issue_records if not r.get("error") and not r["sep_ok"]]
    if sep_anomalies:
        for rec in sep_anomalies:
            nota = (f"AS-IS usa '{rec['sep_asis']}', "
                    f"TO-BE usa '{rec['sep_tobe']}'. "
                    "Il separatore viene rilevato automaticamente per ogni file.")
            _row_colored(ws_sep, [
                rec["label"], rec["asis_label"], rec["tobe_label"],
                rec["sep_asis"], rec["sep_tobe"], nota,
            ], _C_WARN)
    else:
        ws_sep.append(["(nessuna coppia con separatori diversi)"])
    _xl_autofit(ws_sep)

    # ── 3. ANOMALIE_STRUTTURA ─────────────────────────────────────────────────
    ws_str = wb.create_sheet("ANOMALIE_STRUTTURA")
    _xl_title(ws_str, "Colonne presenti in un solo file della coppia", 4)
    ws_str.append(["FILE LOGICO", "COLONNA", "ANOMALIA", "NOTA"])
    _xl_hdr_row(ws_str, 2)

    has_struct = False
    for rec in issue_records:
        if rec.get("error"):
            continue
        for c in rec["cols_only_a"]:
            _row_colored(ws_str, [
                rec["label"], c, "SOLO AS-IS",
                "Colonna presente in AS-IS ma assente in TO-BE",
            ], _C_ONLY_AS)
            has_struct = True
        for c in rec["cols_only_b"]:
            _row_colored(ws_str, [
                rec["label"], c, "SOLO TO-BE",
                "Colonna presente in TO-BE ma assente in AS-IS",
            ], _C_ONLY_TO)
            has_struct = True

    if not has_struct:
        ws_str.append(["(nessuna anomalia strutturale rilevata)"])
    _xl_autofit(ws_str)

    # ── 4. CONFORMITA_TIPI ────────────────────────────────────────────────────
    ws_ti = wb.create_sheet("CONFORMITA_TIPI")
    _xl_title(ws_ti, "Colonne con tipo di dato incoerente tra AS-IS e TO-BE", 5)
    ws_ti.append(["FILE LOGICO", "COLONNA", "TIPO AS-IS", "TIPO TO-BE", "NOTA"])
    _xl_hdr_row(ws_ti, 2)

    has_type = False
    for rec in issue_records:
        if rec.get("error"):
            continue
        for mm in rec["type_mismatches"]:
            _row_colored(ws_ti, [
                rec["label"],
                mm["COLONNA"],
                mm["TIPO AS-IS"],
                mm["TIPO TO-BE"],
                f"Atteso tipo coerente: AS-IS={mm['TIPO AS-IS']}, "
                f"TO-BE={mm['TIPO TO-BE']}",
            ], _C_WARN)
            has_type = True

    if not has_type:
        ws_ti.append(["(nessuna difformita' di tipo rilevata)"])
    _xl_autofit(ws_ti)

    # ── 5. SPAZI ──────────────────────────────────────────────────────────────
    ws_sp = wb.create_sheet("SPAZI")
    _xl_title(ws_sp, "File con campi contenenti spazi da normalizzare", 4)
    ws_sp.append(["FILE LOGICO", "VERSIONE", "FILE", "CAMPI CON SPAZI RILEVATI"])
    _xl_hdr_row(ws_sp, 2)

    has_ws = False
    for rec in issue_records:
        if rec.get("error"):
            continue
        if rec.get("n_ws_a", 0) > 0:
            _row_colored(ws_sp, [
                rec["label"], "AS-IS", rec["asis_label"], rec["n_ws_a"],
            ], _C_WARN)
            has_ws = True
        if rec.get("n_ws_b", 0) > 0:
            _row_colored(ws_sp, [
                rec["label"], "TO-BE", rec["tobe_label"], rec["n_ws_b"],
            ], _C_WARN)
            has_ws = True

    if not has_ws:
        ws_sp.append(["(nessun campo con spazi rilevato)"])

    ws_sp.append([])
    ws_sp.append(["Nota:", "Gli spazi vengono rimossi automaticamente prima del confronto. "
                  "I valori nel report riflettono il dato dopo normalizzazione."])
    _xl_autofit(ws_sp)

    # ── 6. RIGHE_SALTATE ──────────────────────────────────────────────────────
    ws_sk = wb.create_sheet("RIGHE_SALTATE")
    _xl_title(ws_sk, "Righe saltate in lettura (numero campi diverso dall'header)", 5)
    ws_sk.append(["FILE LOGICO", "VERSIONE", "FILE", "N. RIGA (orig.)", "CONTENUTO RIGA"])
    _xl_hdr_row(ws_sk, 2)

    has_bad = False
    for rec in issue_records:
        if rec.get("error"):
            continue
        for line_num, content in rec.get("bad_lines_a", []):
            _row_colored(ws_sk, [
                rec["label"], "AS-IS", rec["asis_label"],
                line_num, content[:500],
            ], _C_WARN)
            has_bad = True
        for line_num, content in rec.get("bad_lines_b", []):
            _row_colored(ws_sk, [
                rec["label"], "TO-BE", rec["tobe_label"],
                line_num, content[:500],
            ], _C_WARN)
            has_bad = True

    if not has_bad:
        ws_sk.append(["(nessuna riga saltata — tutti i file ben formati)"])

    ws_sk.append([])
    ws_sk.append(["Nota:",
                  "Le righe elencate hanno un numero di campi diverso dall'header "
                  "(es. campo contenente il carattere separatore senza virgolette). "
                  "Vengono escluse dal confronto."])
    _xl_autofit(ws_sk)
    # Colonna contenuto riga piu' larga
    ws_sk.column_dimensions["E"].width = 80

    wb.save(output_path)
    return str(output_path)


# ---------------------------------------------------------------------------
# Log errori tecnici
# ---------------------------------------------------------------------------

def log_error(log_path: str | Path, label: str, exc: Exception) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sep = "=" * 70
    msg = (f"\n{sep}\n"
           f"[{ts}] ERRORE su: {label}\n"
           f"Tipo  : {type(exc).__name__}\n"
           f"Motivo: {exc}\n"
           f"--- Traceback ---\n"
           f"{traceback.format_exc()}"
           f"{sep}\n")
    with open(log_path, "a", encoding="utf-8") as fh:
        fh.write(msg)


# ---------------------------------------------------------------------------
# Enumerazione sorgenti (file singolo, ZIP, cartella)
# ---------------------------------------------------------------------------

def enumerate_sources(path: str | Path) -> list[tuple[str, Path, str | None]]:
    """
    Dato un path (file CSV, ZIP, o cartella), restituisce una lista di
    (display_name, zip_or_csv_path, csv_entry_inside_zip_or_None).
    """
    path = Path(path)
    sources = []

    if path.is_dir():
        for f in sorted(path.iterdir()):
            if f.suffix.lower() == ".csv":
                sources.append((f.name, f, None))
            elif f.suffix.lower() == ".zip":
                sources.extend(_enumerate_zip(f))
    elif path.suffix.lower() == ".zip":
        sources.extend(_enumerate_zip(path))
    elif path.suffix.lower() == ".csv":
        sources.append((path.name, path, None))

    return sources


def _enumerate_zip(zip_path: Path) -> list[tuple[str, Path, str]]:
    result = []
    try:
        with zipfile.ZipFile(zip_path) as zf:
            for name in sorted(zf.namelist()):
                if name.lower().endswith(".csv") and not name.startswith("__"):
                    result.append((Path(name).name, zip_path, name))
    except Exception:
        pass
    return result


# ---------------------------------------------------------------------------
# Entry point principale
# ---------------------------------------------------------------------------

def run_comparison(
    asis_path: str | Path,
    tobe_path: str | Path,
    output_dir: str | Path | None = None,
    sep: str | None = None,
    join_key: list[str] | None = None,
    max_diff_rows: int = 10_000,
    progress_cb: Callable[[str], None] | None = None,
    row_filter: dict[str, str] | None = None,
) -> list[str]:
    """
    Confronta tutte le coppie di file trovate in asis_path e tobe_path.
    Restituisce la lista degli Excel generati.
    """
    def _log(msg: str):
        if progress_cb:
            progress_cb(msg)
        else:
            print(msg)

    asis_path = Path(asis_path)
    tobe_path = Path(tobe_path)

    if output_dir is None:
        output_dir = asis_path.parent if asis_path.is_file() else asis_path
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = output_dir / f"flowcheck_errori_{ts}.log"

    if join_key:
        _log(f"Chiave di join manuale: {', '.join(join_key)}")
    else:
        _log("Chiave di join: auto-detect")

    asis_sources = enumerate_sources(asis_path)
    tobe_sources = enumerate_sources(tobe_path)

    _log(f"AS-IS: {len(asis_sources)} file trovati in {asis_path.name}")
    _log(f"TO-BE: {len(tobe_sources)} file trovati in {tobe_path.name}")

    # Costruisci lista per matching
    asis_list = [(dn, str(zp) if ze is None else f"{zp}::{ze}") for dn, zp, ze in asis_sources]
    tobe_list = [(dn, str(zp) if ze is None else f"{zp}::{ze}") for dn, zp, ze in tobe_sources]

    pairs = match_files(asis_list, tobe_list)
    total_pairs = sum(1 for p in pairs if p["tobe_path"])
    _log(f"Coppie abbinate: {total_pairs}/{len(pairs)}")
    _log("")

    generated    = []
    issue_records: list[dict] = []

    import time as _time

    def _fmt_time(secs: float) -> str:
        secs = max(0.0, secs)
        if secs < 60:
            return f"{secs:.0f}s"
        m, s = divmod(int(secs), 60)
        return f"{m}m {s:02d}s"

    t0   = _time.perf_counter()
    done = 0

    for pair in pairs:
        label = pair["label"]
        if pair["tobe_path"] is None:
            _log(f"[SKIP] {label} - nessuna controparte TO-BE trovata")
            continue

        t_file_start = _time.perf_counter()
        _log(f"[FILE_START] {done + 1}/{total_pairs} {label}")
        _log(f"  Confronto: {label} ...")

        try:
            # ── Rileva separatori (per issue log) ──────────────────────────
            a_path_str = pair["asis_path"]
            b_path_str = pair["tobe_path"]

            # ── Controlla dimensione: modalità streaming per file grandi ───
            size_a   = _source_size(a_path_str)
            size_b   = _source_size(b_path_str)
            is_large = max(size_a, size_b) >= _LARGE_FILE_THRESHOLD

            if "::" in a_path_str:
                zp_a, ze_a = a_path_str.split("::", 1)
                sep_a = sep if sep else detect_separator(zp_a, zip_entry=ze_a)
            else:
                sep_a = sep if sep else detect_separator(a_path_str)

            if "::" in b_path_str:
                zp_b, ze_b = b_path_str.split("::", 1)
                sep_b = sep if sep else detect_separator(zp_b, zip_entry=ze_b)
            else:
                sep_b = sep if sep else detect_separator(b_path_str)

            # ── Output path condiviso ──────────────────────────────────────
            out_name = f"confronto_{label}_{ts}.xlsx"
            out_path = output_dir / out_name

            if is_large:
                # ── Modalità STREAMING (SQLite) ────────────────────────────
                mb = max(size_a, size_b) / 1024 / 1024
                _log(f"  [LARGE] File grande ({mb:.0f} MB) — modalità streaming SQLite")

                generated_path = build_excel_pair_large(
                    label=label,
                    path_a=a_path_str,
                    path_b=b_path_str,
                    sep_a=sep_a,
                    sep_b=sep_b,
                    asis_label=pair["asis_label"],
                    tobe_label=pair["tobe_label"],
                    output_path=out_path,
                    join_key=join_key,
                    log_cb=_log,
                    row_filter=row_filter,
                )
                generated.append(generated_path)

                # Issue log per modalità streaming (campi richiesti con valori default)
                issue_records.append({
                    "label":           label,
                    "asis_label":      pair["asis_label"],
                    "tobe_label":      pair["tobe_label"],
                    "sep_a":           sep_a,
                    "sep_b":           sep_b,
                    "sep_ok":          sep_a == sep_b,
                    "n_a":             0,
                    "n_b":             0,
                    "cols_only_a":     [],
                    "cols_only_b":     [],
                    "type_mismatches": [],
                    "n_diff_rows":     0,
                    "n_only_a":        0,
                    "n_only_b":        0,
                    "bad_lines_a":     [],
                    "bad_lines_b":     [],
                    "n_ws_a":          0,
                    "n_ws_b":          0,
                    "note":            f"Streaming DuckDB ({mb:.0f} MB) — dettagli nel file Excel",
                })

                _log(f"  [OK] Excel salvato: {out_name}")

            else:
                # ── Modalità NORMALE (in-memory DataFrame) ─────────────────
                if "::" in a_path_str:
                    df_a   = read_csv_from_zip(zp_a, ze_a, sep=sep)
                    raw_a  = _get_raw_content(zp_a, ze_a)
                    n_ws_a = _detect_whitespace_in_raw(zp_a, ze_a, sep_a)
                else:
                    df_a   = read_csv(a_path_str, sep=sep)
                    raw_a  = _get_raw_content(a_path_str)
                    n_ws_a = _detect_whitespace_in_raw(a_path_str, None, sep_a)

                if "::" in b_path_str:
                    df_b   = read_csv_from_zip(zp_b, ze_b, sep=sep)
                    raw_b  = _get_raw_content(zp_b, ze_b)
                    n_ws_b = _detect_whitespace_in_raw(zp_b, ze_b, sep_b)
                else:
                    df_b   = read_csv(b_path_str, sep=sep)
                    raw_b  = _get_raw_content(b_path_str)
                    n_ws_b = _detect_whitespace_in_raw(b_path_str, None, sep_b)

                # ── Applica filtro righe se specificato ───────────────────
                if row_filter:
                    for _fc, _fv in row_filter.items():
                        if _fc in df_a.columns:
                            df_a = df_a[df_a[_fc].str.strip() == _fv.strip()].reset_index(drop=True)
                        if _fc in df_b.columns:
                            df_b = df_b[df_b[_fc].str.strip() == _fv.strip()].reset_index(drop=True)
                    _filter_desc = "  AND  ".join(f"{k} = '{v}'" for k, v in row_filter.items())
                    _log(f"  Filtro righe: {_filter_desc}  →  AS-IS {len(df_a):,} righe, TO-BE {len(df_b):,} righe")

                # ── Rileva righe saltate ───────────────────────────────────
                bad_lines_a = _find_bad_lines(raw_a, sep_a)
                bad_lines_b = _find_bad_lines(raw_b, sep_b)

                if bad_lines_a:
                    _log(f"  [ATTENZIONE] AS-IS {pair['asis_label']}: "
                         f"{len(bad_lines_a)} righe saltate (campi non conformi)")
                    for ln, content in bad_lines_a[:5]:
                        _log(f"    riga {ln}: {content[:120]}")
                    if len(bad_lines_a) > 5:
                        _log(f"    ... e altre {len(bad_lines_a) - 5} righe")

                if bad_lines_b:
                    _log(f"  [ATTENZIONE] TO-BE {pair['tobe_label']}: "
                         f"{len(bad_lines_b)} righe saltate (campi non conformi)")
                    for ln, content in bad_lines_b[:5]:
                        _log(f"    riga {ln}: {content[:120]}")
                    if len(bad_lines_b) > 5:
                        _log(f"    ... e altre {len(bad_lines_b) - 5} righe")

                # ── Genera Excel per coppia ────────────────────────────────
                _filter_desc = ("  AND  ".join(f"{k} = '{v}'" for k, v in row_filter.items())
                                if row_filter else None)
                generated_path = build_excel_pair(
                    label=label,
                    df_a=df_a,
                    df_b=df_b,
                    asis_label=pair["asis_label"],
                    tobe_label=pair["tobe_label"],
                    output_path=out_path,
                    join_key=join_key,
                    max_diff_rows=max_diff_rows,
                    log_path=log_path,
                    row_filter_desc=_filter_desc,
                )
                generated.append(generated_path)

                # ── Raccoglie dati per issue log ───────────────────────────
                issue_records.append(_collect_pair_issues(
                    label=label,
                    asis_label=pair["asis_label"],
                    tobe_label=pair["tobe_label"],
                    df_a=df_a, df_b=df_b,
                    sep_a=sep_a, sep_b=sep_b,
                    n_ws_a=n_ws_a, n_ws_b=n_ws_b,
                    bad_lines_a=bad_lines_a,
                    bad_lines_b=bad_lines_b,
                ))

                _log(f"  [OK] Excel salvato: {out_name}")

        except Exception as exc:
            log_error(log_path, label, exc)
            _log(f"  [ERRORE] {label} - vedi log: {log_path.name}")
            issue_records.append({
                "label": label,
                "asis_label": pair.get("asis_label", ""),
                "tobe_label": pair.get("tobe_label", ""),
                "error": str(exc),
            })

        # ── Timing e avanzamento ───────────────────────────────────────
        done += 1
        t_file   = _time.perf_counter() - t_file_start
        t_total  = _time.perf_counter() - t0
        t_avg    = t_total / done
        t_eta    = t_avg * (total_pairs - done)
        timing   = f"  ⏱  File: {_fmt_time(t_file)}"
        timing  += f"  |  Trascorso: {_fmt_time(t_total)}"
        if total_pairs - done > 0:
            timing += f"  |  Stima rimanente: ~{_fmt_time(t_eta)}"
        _log(timing)
        # Messaggio strutturato per la barra di avanzamento nella GUI
        _log(f"[PROGRESS] {done}/{total_pairs}")

    # ── Genera issue log Excel ─────────────────────────────────────────────
    if issue_records:
        issue_log_path = output_dir / f"issue_log_{ts}.xlsx"
        try:
            build_issue_log(issue_records, issue_log_path)
            _log(f"Issue log: {issue_log_path.name}")
        except Exception as exc:
            log_error(log_path, "build_issue_log", exc)
            _log(f"  [ATTENZIONE] Issue log non generato: {exc}")

    _log("")
    _log(f"Completato. {len(generated)}/{len([p for p in pairs if p['tobe_path']])} Excel generati.")
    if log_path.exists():
        _log(f"Log errori tecnici: {log_path.name}")
    return generated


# ---------------------------------------------------------------------------
# CLI minimale (test rapido)
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Uso: python flowcheck_engine.py <AS-IS> <TO-BE> [output_dir]")
        sys.exit(1)
    out = sys.argv[3] if len(sys.argv) > 3 else None
    run_comparison(sys.argv[1], sys.argv[2], output_dir=out)
